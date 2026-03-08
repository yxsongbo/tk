#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据导入脚本
1. 从 student.xlsx 导入学生数据
2. 从 tk202601.json 导入题目数据
"""

import json
import sqlite3
from openpyxl import load_workbook


def _to_str_list(value):
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    text = str(value).strip()
    return [text] if text else []


def _to_image_list(value):
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return []
        return [text]
    text = str(value).strip()
    return [text] if text else []


def _extract_fill_correct_answer(question):
    answers = question.get("answers")
    if not isinstance(answers, list):
        return None

    blanks = []
    for item in answers:
        if isinstance(item, dict):
            blank_answers = _to_str_list(item.get("answers"))
            blank_images = _to_image_list(item.get("images"))
            if not blank_images:
                blank_images = _to_image_list(item.get("image"))
            label = item.get("label")
            blank_label = str(label).strip() if label is not None else None
            if blank_label == "":
                blank_label = None
            if blank_answers or blank_images or blank_label:
                blanks.append(
                    {"answers": blank_answers, "images": blank_images, "label": blank_label}
                )
        elif isinstance(item, list):
            values = _to_str_list(item)
            if values:
                blanks.append({"answers": values, "images": [], "label": None})
        else:
            value = str(item).strip()
            if value:
                blanks.append({"answers": [value], "images": [], "label": None})

    return json.dumps(blanks, ensure_ascii=False) if blanks else None


def import_students():
    """导入学生数据"""
    print("正在导入学生数据...")

    # 读取Excel（不使用pandas，使用openpyxl）
    wb = load_workbook("student.xlsx")
    ws = wb.active

    conn = sqlite3.connect("practice.db")
    cursor = conn.cursor()

    # 清空现有数据
    cursor.execute("DELETE FROM students")

    # 跳过表头，从第二行开始
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:  # 跳过空行
            continue

        exam_number = str(row[0])
        class_number = int(row[1]) if row[1] else 0
        student_number = int(row[2]) if row[2] else 0
        name = str(row[3])
        subject_group = int(row[4]) if row[4] else 0

        cursor.execute(
            """
            INSERT INTO students (exam_number, class_number, student_number, name, subject_group)
            VALUES (?, ?, ?, ?, ?)
            """,
            (exam_number, class_number, student_number, name, subject_group),
        )
        count += 1

    conn.commit()
    conn.close()
    print(f"成功导入 {count} 名学生")


def find_json_files():
    """查找当前目录下的所有JSON文件"""
    from pathlib import Path

    base_dir = Path(".")
    exams_dir = base_dir / "exams"
    files = {}
    for p in base_dir.glob("*.json"):
        files[p.name] = p
    if exams_dir.exists():
        for p in exams_dir.glob("*.json"):
            files[p.name] = p
    return [str(files[name]) for name in sorted(files.keys())]


def select_json_file(json_files):
    """让用户选择要导入的JSON文件"""
    if not json_files:
        return None
    if len(json_files) == 1:
        print(f"自动选择唯一文件: {json_files[0]}")
        return json_files[0]

    print("\n可用试卷文件:")
    for i, f in enumerate(json_files):
        print(f"  {i + 1}. {f}")

    while True:
        try:
            choice = input("\n请选择要导入的试卷编号 (1-{}): ".format(len(json_files)))
            idx = int(choice) - 1
            if 0 <= idx < len(json_files):
                return json_files[idx]
        except ValueError:
            pass
        print("无效选择，请重试")


def import_questions():
    """导入题目数据"""
    print("正在导入题目数据...")

    # 查找所有JSON文件
    json_files = find_json_files()
    if not json_files:
        print("错误: 未找到任何JSON文件")
        return

    # 选择要导入的文件
    selected_file = select_json_file(json_files)
    if not selected_file:
        print("未选择文件")
        return

    print(f"\n正在从 {selected_file} 导入...")

    # 读取JSON
    try:
        with open(selected_file, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        print(f"错误: 读取文件失败 - {e}")
        return

    conn = sqlite3.connect("practice.db")
    cursor = conn.cursor()

    # 清空现有数据
    cursor.execute("DELETE FROM questions")

    # 插入数据
    count = 0
    for q in data["questions"]:
        options = (
            json.dumps(q.get("options", []), ensure_ascii=False)
            if "options" in q
            else None
        )
        image_value = q.get("image")
        if not image_value and isinstance(q.get("images"), list):
            image_value = json.dumps(q.get("images", []), ensure_ascii=False)

        # 处理答案字段: 优先使用 correctAnswer，其次使用 answers
        if q.get("type") == "fill":
            correct_answer = _extract_fill_correct_answer(q)
            if correct_answer is None:
                correct_answer = q.get("correctAnswer")
        else:
            correct_answer = q.get("correctAnswer")
            if correct_answer is None and "answers" in q:
                answers = q.get("answers", [])
                if answers:
                    all_answers = []
                    for answer_group in answers:
                        if isinstance(answer_group, list):
                            all_answers.extend([str(a) for a in answer_group if a])
                        elif answer_group:
                            all_answers.append(str(answer_group))
                    correct_answer = ",".join(all_answers)

        cursor.execute(
            """
            INSERT INTO questions (id, type, question, score, explanation, image, correct_answer, options)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                q["id"],
                q["type"],
                q["question"],
                q["score"],
                q.get("explanation", ""),
                image_value,
                correct_answer,
                options,
            ),
        )
        count += 1

    conn.commit()
    conn.close()
    print(f"成功导入 {count} 道题目")


def check_data():
    """检查导入的数据"""
    conn = sqlite3.connect("practice.db")
    cursor = conn.cursor()

    print("\n=== 数据检查 ===")

    # 学生数量
    cursor.execute("SELECT COUNT(*) FROM students")
    print(f"学生总数: {cursor.fetchone()[0]}")

    # 班级分布
    cursor.execute("SELECT class_number, COUNT(*) FROM students GROUP BY class_number")
    print("\n班级分布:")
    for row in cursor.fetchall():
        print(f"  {row[0]}班: {row[1]}人")

    # 题目数量
    cursor.execute("SELECT COUNT(*) FROM questions")
    print(f"\n题目总数: {cursor.fetchone()[0]}")

    # 题目类型分布
    cursor.execute("SELECT type, COUNT(*) FROM questions GROUP BY type")
    print("\n题目类型分布:")
    for row in cursor.fetchall():
        print(f"  {row[0]}: {row[1]}道")

    conn.close()


if __name__ == "__main__":
    import_students()
    import_questions()
    check_data()
    print("\n数据导入完成！")
