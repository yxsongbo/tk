#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
练习系统后端API
支持100人同时在线练习
技术栈: FastAPI + SQLite
"""

import json
import ast
import re
import sqlite3
import os
import base64
import hashlib
from datetime import datetime, timedelta
from typing import Any, List, Optional
from contextlib import asynccontextmanager
from pathlib import Path
import subprocess
import threading
import queue
import time

from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, RedirectResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import io
from urllib.parse import quote
from pydantic import BaseModel, Field
import uvicorn

# 项目路径
BASE_DIR = Path(__file__).resolve().parent.parent
APP_DIR = BASE_DIR / "app"
EXAMS_DIR = BASE_DIR / "exams"
GENERATED_IMAGES_DIR = BASE_DIR / "generated_images"

# 数据库路径
DATABASE = str(BASE_DIR / "practice.db")
DB_BACKEND = "postgres" if os.getenv("DB_BACKEND") == "postgres" else "sqlite"
DATABASE_URL = os.getenv("DATABASE_URL", "")
if DATABASE_URL.startswith("postgres"):
    DB_BACKEND = "postgres"

try:
    import psycopg
    from psycopg.rows import dict_row
    from psycopg_pool import ConnectionPool
except Exception:  # pragma: no cover - optional dependency for Postgres
    psycopg = None
    dict_row = None
    ConnectionPool = None

PG_POOL: Optional["ConnectionPool"] = None
ANSWER_QUEUE: Optional["queue.Queue"] = None
ANSWER_WORKER: Optional[threading.Thread] = None
ANSWER_STOP = threading.Event()
SIDEBAR_HISTORY_CACHE: dict[str, dict[str, Any]] = {}
SIDEBAR_HISTORY_CACHE_LOCK = threading.Lock()
SIDEBAR_HISTORY_TTL_SECONDS = 3600
SIDEBAR_ACTIVE_LIMIT = 8


def _is_postgres() -> bool:
    return DB_BACKEND == "postgres"


def _pg_conninfo() -> str:
    if DATABASE_URL:
        return DATABASE_URL
    host = os.getenv("PGHOST", "127.0.0.1")
    port = os.getenv("PGPORT", "5432")
    dbname = os.getenv("PGDATABASE", "tk_practice")
    user = os.getenv("PGUSER", "songbo")
    password = os.getenv("PGPASSWORD", "")
    return (
        f"host={host} port={port} dbname={dbname} user={user} password={password}"
    )


def _fetchone_value(row, default=None):
    if row is None:
        return default
    if isinstance(row, dict):
        return next(iter(row.values()))
    return row[0]


def _start_answer_worker() -> None:
    if not _is_postgres():
        return
    global ANSWER_QUEUE, ANSWER_WORKER, PG_POOL
    if PG_POOL is None:
        if ConnectionPool is None:
            raise RuntimeError("psycopg_pool 未安装，无法创建连接池")
        min_size = int(os.getenv("PG_POOL_MIN", "5"))
        max_size = int(os.getenv("PG_POOL_MAX", "60"))
        timeout = float(os.getenv("PG_POOL_TIMEOUT", "10"))
        PG_POOL = ConnectionPool(
            _pg_conninfo(),
            min_size=min_size,
            max_size=max_size,
            timeout=timeout,
            kwargs={"row_factory": dict_row},
        )
    if ANSWER_QUEUE is None:
        ANSWER_QUEUE = queue.Queue(maxsize=5000)
    if ANSWER_WORKER is None or not ANSWER_WORKER.is_alive():
        ANSWER_STOP.clear()
        ANSWER_WORKER = threading.Thread(target=_answer_worker_loop, daemon=True)
        ANSWER_WORKER.start()


def _stop_answer_worker() -> None:
    global ANSWER_WORKER
    if ANSWER_WORKER is None:
        return
    ANSWER_STOP.set()
    ANSWER_WORKER.join(timeout=3)


def _flush_answers(batch: list[tuple]) -> None:
    if not batch:
        return
    if PG_POOL is None:
        return
    with PG_POOL.connection() as conn:
        cursor = conn.cursor()
        cursor.executemany(
            """
            INSERT INTO answers
            (session_id, student_id, question_id, answer, is_correct, score, answer_time, created_at, exam_filename)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT(session_id, question_id) DO UPDATE SET
                answer = excluded.answer,
                is_correct = excluded.is_correct,
                score = excluded.score,
                answer_time = excluded.answer_time,
                created_at = excluded.created_at,
                exam_filename = excluded.exam_filename
            """,
            batch,
        )
        conn.commit()


def _answer_worker_loop() -> None:
    batch: list[tuple] = []
    last_flush = time.time()
    while not ANSWER_STOP.is_set():
        try:
            item = ANSWER_QUEUE.get(timeout=0.2) if ANSWER_QUEUE else None
            if item:
                batch.append(item)
        except queue.Empty:
            pass

        now = time.time()
        if batch and (len(batch) >= 50 or now - last_flush >= 0.5):
            try:
                _flush_answers(batch)
            except Exception:
                # Fallback: try one by one to avoid losing data
                for single in batch:
                    try:
                        _flush_answers([single])
                    except Exception:
                        pass
            batch = []
            last_flush = now

    # Drain remaining on shutdown
    if batch:
        try:
            _flush_answers(batch)
        except Exception:
            pass


class _SQLiteCompatCursor:
    def __init__(self, cursor: sqlite3.Cursor):
        self._cursor = cursor

    def execute(self, sql: str, params=()):
        return self._cursor.execute(sql.replace("%s", "?"), params)

    def executemany(self, sql: str, seq_of_params):
        return self._cursor.executemany(sql.replace("%s", "?"), seq_of_params)

    def __getattr__(self, name: str):
        return getattr(self._cursor, name)


class _SQLiteCompatConn:
    def __init__(self, conn: sqlite3.Connection):
        self._conn = conn

    def cursor(self):
        return _SQLiteCompatCursor(self._conn.cursor())

    def __getattr__(self, name: str):
        return getattr(self._conn, name)


def normalize_fill_answer(answer: str) -> str:
    """统一填空答案格式：去首尾空白、去全角空格、忽略中间空格。"""
    return str(answer).strip().replace("\u3000", "").replace(" ", "")


def levenshtein_distance(a: str, b: str) -> int:
    """计算两个字符串之间的编辑距离。"""
    if len(a) == 0:
        return len(b)
    if len(b) == 0:
        return len(a)

    matrix = [[0] * (len(b) + 1) for _ in range(len(a) + 1)]

    for i in range(len(a) + 1):
        matrix[i][0] = i
    for j in range(len(b) + 1):
        matrix[0][j] = j

    for i in range(1, len(a) + 1):
        for j in range(1, len(b) + 1):
            if a[i - 1] == b[j - 1]:
                matrix[i][j] = matrix[i - 1][j - 1]
            else:
                matrix[i][j] = min(
                    matrix[i - 1][j - 1] + 1,  # substitution
                    matrix[i][j - 1] + 1,  # insertion
                    matrix[i - 1][j] + 1,  # deletion
                )

    return matrix[len(a)][len(b)]


def strings_similarity(a: str, b: str) -> float:
    """计算两个字符串的相似度（0.0-1.0）。"""
    if a == b:
        return 1.0
    if len(a) == 0 or len(b) == 0:
        return 0.0

    distance = levenshtein_distance(a, b)
    max_length = max(len(a), len(b))
    return 1.0 - (distance / max_length)


def build_fill_answer_variants(answer: str) -> set[str]:
    """从标准答案中提取可接受答案（支持“或”前后写法）。"""
    raw = str(answer).strip()
    if not raw:
        return set()

    variants = {normalize_fill_answer(raw)}

    normalized_brackets = raw.replace("（", "(").replace("）", ")")

    # 兼容：A（或B）/ A(或B) / A或B / A，或B / A,或B
    if "或" in normalized_brackets:
        parts = re.split(r"\s*[,，]?\s*或\s*", normalized_brackets)
        for part in parts:
            cleaned = part.strip().strip("()[]{}<>\"'“”‘’、，,;；")
            candidate = normalize_fill_answer(cleaned)
            if candidate:
                variants.add(candidate)

            # 兼容：字符串(str) 这种混合写法
            m = re.match(r"^(.*?)\((.*?)\)$", cleaned)
            if m:
                left = normalize_fill_answer(m.group(1))
                right = normalize_fill_answer(m.group(2))
                if left:
                    variants.add(left)
                if right:
                    variants.add(right)

    return {v for v in variants if v}


def is_fill_answer_match(student_answer: str, correct_answer: str) -> float:
    """填空匹配：含“或”时，允许匹配“或”前后内容（含近似包含）。返回0.0-1.0的相似度，>=0.9返回1.0（满分）。"""
    student = normalize_fill_answer(student_answer)
    if not student:
        return 0.0

    raw = str(correct_answer or "").strip()
    variants = build_fill_answer_variants(raw)
    if not variants:
        return 0.0

    max_similarity = 0.0

    # Check exact match first
    if student in variants:
        return 1.0

    # Calculate similarity with all variants
    for variant in variants:
        similarity = strings_similarity(student, variant)
        if similarity > max_similarity:
            max_similarity = similarity

    # For answers containing "或", also check substring relationship
    if "或" in raw.replace("（", "(").replace("）", ")"):
        for variant in variants:
            if (
                len(student) >= 2
                and len(variant) >= 2
                and (student in variant or variant in student)
            ):
                # Calculate similarity for substring relationship
                shorter = student if len(student) < len(variant) else variant
                longer = student if len(student) >= len(variant) else variant
                similarity = strings_similarity(shorter, longer)
                if similarity > max_similarity:
                    max_similarity = similarity

    # If similarity >= 90%, return full score
    if max_similarity >= 0.9:
        return 1.0

    return max_similarity


def _coerce_str_list(value: Any) -> list[str]:
    """把值统一转换为字符串数组（去掉空值）。"""
    if value is None:
        return []
    if isinstance(value, list):
        values = []
        for item in value:
            if item is None:
                continue
            text = str(item).strip()
            if text:
                values.append(text)
        return values
    text = str(value).strip()
    return [text] if text else []


FORMULA_FRAGMENT_RE = re.compile(r"(=|\$|\(|\)|:|[A-Za-z]+\$?\d+)")


def _dedupe_preserve_order(items: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        if item not in seen:
            seen.add(item)
            result.append(item)
    return result


def _should_join_answer_fragments(answers: list[str]) -> bool:
    if len(answers) < 2:
        return False
    if any("或" in answer for answer in answers):
        return False
    if any("," in answer or "，" in answer for answer in answers):
        return False
    combined = ",".join(answer.strip() for answer in answers if answer)
    if not combined:
        return False
    if combined.startswith("="):
        return True
    return bool(FORMULA_FRAGMENT_RE.search(combined))


def _expand_fill_blank_answers(answers: list[str]) -> list[str]:
    cleaned = [str(item).strip() for item in answers if str(item).strip()]
    if not cleaned:
        return []
    cleaned = _dedupe_preserve_order(cleaned)
    if _should_join_answer_fragments(cleaned):
        for joiner in (",", "，"):
            joined = joiner.join(cleaned)
            if joined and joined not in cleaned:
                cleaned.append(joined)
    return cleaned


def _coerce_image_list(value: Any) -> list[str]:
    """把单图/多图字段统一为字符串数组。"""
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return []
        if raw.startswith("["):
            try:
                parsed = json.loads(raw)
                if isinstance(parsed, list):
                    return [str(v).strip() for v in parsed if str(v).strip()]
            except Exception:
                pass
        return [raw]
    text = str(value).strip()
    return [text] if text else []


def _answer_to_choice_index(answer: Any) -> int | str | None:
    """把旧格式选择题答案统一为零基索引。"""
    if answer is None:
        return None
    if isinstance(answer, int):
        return answer

    text = str(answer).strip()
    if not text:
        return None
    if text.isdigit():
        return int(text)

    normalized = text.upper()
    if len(normalized) == 1 and "A" <= normalized <= "Z":
        return ord(normalized) - ord("A")
    return text


def default_score_for_type(question_type: str) -> float:
    question_type = str(question_type or "").strip()
    if question_type == "fill":
        return 10.0
    return 2.5


def _normalize_legacy_choice_question(question: dict, index: int) -> dict:
    """兼容旧格式选择题。"""
    options = []
    for option in question.get("options", []):
        if isinstance(option, dict):
            options.append(
                {
                    "text": str(option.get("text", "")).strip(),
                    "image": option.get("image"),
                }
            )
        else:
            options.append({"text": str(option).strip(), "image": None})

    images = _coerce_image_list(question.get("images"))
    if not images:
        images = _coerce_image_list(question.get("imageData"))
    if not images:
        images = _coerce_image_list(question.get("image"))

    normalized_question = {
        "id": str(question.get("id") or f"choice-{index + 1}"),
        "type": "choice",
        "question": str(question.get("question") or question.get("text") or "").strip(),
        "score": float(question.get("score") or default_score_for_type("choice")),
        "explanation": str(question.get("explanation") or "").strip(),
        "options": options,
        "correctAnswer": _answer_to_choice_index(
            question.get("correctAnswer", question.get("answer"))
        ),
    }
    if images:
        normalized_question["images"] = images
    return normalized_question


def _normalize_legacy_fill_question(question: dict, index: int) -> dict:
    """兼容旧格式填空题。"""
    blanks = []
    for blank in question.get("blanks", []):
        if not isinstance(blank, dict):
            continue
        answers = _expand_fill_blank_answers(
            _coerce_str_list(blank.get("correctAnswers", blank.get("answers")))
        )
        label = blank.get("label")
        label_text = str(label).strip() if label is not None else None
        if label_text == "":
            label_text = None
        images = _coerce_image_list(blank.get("images"))
        if not images:
            images = _coerce_image_list(blank.get("image"))
        blanks.append({"answers": answers, "images": images, "label": label_text})

    images = _coerce_image_list(question.get("images"))
    if not images:
        images = _coerce_image_list(question.get("imageData"))
    if not images:
        images = _coerce_image_list(question.get("image"))

    normalized_question = {
        "id": str(question.get("id") or f"fill-{index + 1}"),
        "type": "fill",
        "question": str(question.get("question") or question.get("text") or "").strip(),
        "score": float(question.get("score") or default_score_for_type("fill")),
        "explanation": str(question.get("explanation") or "").strip(),
        "answers": blanks,
    }
    if images:
        normalized_question["images"] = images
    return normalized_question


def normalize_exam_json(data: dict) -> dict:
    """把历史试卷格式转换为当前统一格式。"""
    questions = data.get("questions")
    if isinstance(questions, list) and questions:
        return data

    choice_questions = data.get("choiceQuestions")
    fill_questions = data.get("fillQuestions")
    if not isinstance(choice_questions, list) and not isinstance(fill_questions, list):
        return data

    normalized_questions = []
    if isinstance(choice_questions, list):
        for index, question in enumerate(choice_questions):
            if isinstance(question, dict):
                normalized_questions.append(
                    _normalize_legacy_choice_question(question, index)
                )

    if isinstance(fill_questions, list):
        base_index = len(normalized_questions)
        for index, question in enumerate(fill_questions):
            if isinstance(question, dict):
                normalized_questions.append(
                    _normalize_legacy_fill_question(question, base_index + index)
                )

    normalized_data = dict(data)
    normalized_data["questions"] = normalized_questions
    if "title" not in normalized_data and data.get("name"):
        normalized_data["title"] = str(data.get("name")).strip()
    return normalized_data


def _normalize_fill_blank_item(item: Any) -> dict:
    """标准化填空每一空的结构。"""
    if isinstance(item, dict):
        answers = _expand_fill_blank_answers(_coerce_str_list(item.get("answers")))
        images = _coerce_image_list(item.get("images"))
        if not images:
            images = _coerce_image_list(item.get("image"))
        label = item.get("label")
        label_text = str(label).strip() if label is not None else None
        if label_text == "":
            label_text = None
        return {"answers": answers, "images": images, "label": label_text}

    if isinstance(item, list):
        return {
            "answers": _expand_fill_blank_answers(_coerce_str_list(item)),
            "images": [],
            "label": None,
        }

    return {
        "answers": _expand_fill_blank_answers(_coerce_str_list(item)),
        "images": [],
        "label": None,
    }


def parse_fill_blanks(raw: Any) -> list[dict]:
    """
    解析填空答案，兼容多种历史格式：
    1) JSON数组/对象字符串
    2) Python 字面量字符串（旧数据）
    3) 逗号分隔字符串
    """
    if raw is None:
        return []

    parsed: Any = raw
    if isinstance(raw, str):
        text = raw.strip()
        if not text:
            return []

        parsed = text
        parse_candidates: list[str] = []
        if text.startswith("[") or text.startswith("{"):
            parse_candidates.append(text)
        parse_candidates.append(text)
        parse_candidates.append(f"[{text}]")

        for candidate in parse_candidates:
            try:
                parsed = json.loads(candidate)
                break
            except Exception:
                try:
                    parsed = ast.literal_eval(candidate)
                    break
                except Exception:
                    parsed = text

        if isinstance(parsed, str):
            # Treat the entire string as a single acceptable answer for one blank
            # Don't split on commas, as they may be part of the answer (e.g., Excel formulas)
            return [{"answers": [text], "images": [], "label": None}]

    if isinstance(parsed, dict):
        if isinstance(parsed.get("blanks"), list):
            parsed = parsed.get("blanks")
        elif isinstance(parsed.get("answers"), list):
            parsed = parsed.get("answers")
        else:
            return []

    if isinstance(parsed, list):
        blanks = [_normalize_fill_blank_item(item) for item in parsed]
        return [
            blank
            for blank in blanks
            if blank["answers"] or blank["images"] or blank["label"]
        ]

    return [_normalize_fill_blank_item(parsed)]


def ensure_column(cursor: Any, table: str, column: str, ddl: str) -> None:
    """为已有表补充字段（若不存在）。"""
    if _is_postgres():
        cursor.execute(
            "SELECT column_name FROM information_schema.columns WHERE table_name = %s",
            (table,),
        )
        rows = cursor.fetchall()
        existing_columns = {
            (row["column_name"] if isinstance(row, dict) else row[0]) for row in rows
        }
    else:
        cursor.execute(f"PRAGMA table_info({table})")
        existing_columns = {row[1] for row in cursor.fetchall()}
    if column not in existing_columns:
        cursor.execute(f"ALTER TABLE {table} ADD COLUMN {ddl}")


def ensure_exam_questions_table(cursor: sqlite3.Cursor) -> None:
    """确保 exam_questions 表存在，用于缓存多份试卷的题目。"""
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS exam_questions (
            exam_filename TEXT NOT NULL,
            id TEXT NOT NULL,
            type TEXT NOT NULL CHECK(type IN ('choice', 'fill')),
            question TEXT NOT NULL,
            score REAL NOT NULL,
            explanation TEXT,
            image TEXT,
            correct_answer TEXT,
            options TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (exam_filename, id)
        )
        """
    )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_exam_questions_exam ON exam_questions(exam_filename)"
    )


def read_meminfo() -> dict[str, int]:
    """读取 /proc/meminfo，返回 kB 数值字典；失败则返回空。"""
    info: dict[str, int] = {}
    try:
        with open("/proc/meminfo", "r") as f:
            for line in f:
                parts = line.split(":")
                if len(parts) >= 2:
                    key = parts[0].strip()
                    value = parts[1].strip().split()[0]
                    info[key] = int(value)
    except Exception:
        pass
    return info


def read_swaps() -> tuple[int, int]:
    """读取 /proc/swaps，返回 (total_kb, used_kb)。"""
    total = used = 0
    try:
        with open("/proc/swaps", "r") as f:
            lines = f.readlines()[1:]  # skip header
            for line in lines:
                parts = line.split()
                if len(parts) >= 5:
                    total += int(parts[2])
                    used += int(parts[3])
    except Exception:
        pass
    return total, used


def get_system_health() -> dict:
    """轻量级系统负载指标，避免额外依赖。"""
    meminfo = read_meminfo()
    mem_avail_kb = meminfo.get("MemAvailable", 0)
    mem_total_kb = meminfo.get("MemTotal", 0)
    swap_total_kb, swap_used_kb = read_swaps()
    load1, load5, load15 = os.getloadavg() if hasattr(os, "getloadavg") else (0, 0, 0)

    # 简单告警判定：可用内存 < 80MB 或 swap 使用率 > 80% 或 1分钟负载 > CPU核数*1.5
    cpu_count = os.cpu_count() or 1
    mem_alert = mem_avail_kb < 80 * 1024
    swap_alert = swap_total_kb > 0 and (swap_used_kb / swap_total_kb) > 0.8
    load_alert = load1 > cpu_count * 1.5
    degraded = mem_alert or swap_alert or load_alert

    return {
        "mem_available_mb": round(mem_avail_kb / 1024, 1),
        "mem_total_mb": round(mem_total_kb / 1024, 1),
        "swap_total_mb": round(swap_total_kb / 1024, 1),
        "swap_used_mb": round(swap_used_kb / 1024, 1),
        "load1": round(load1, 2),
        "load5": round(load5, 2),
        "load15": round(load15, 2),
        "cpu_count": cpu_count,
        "alert": degraded,
        "alert_reasons": {
            "low_memory": mem_alert,
            "high_swap": swap_alert,
            "high_load": load_alert,
        },
    }


def get_sidebar_online_refresh_ms(health: dict) -> int:
    """根据当前负载给出在线榜单建议刷新间隔。"""
    if health.get("alert"):
        return 45000

    load1 = float(health.get("load1") or 0)
    cpu_count = max(int(health.get("cpu_count") or 1), 1)
    load_ratio = load1 / cpu_count
    mem_available_mb = float(health.get("mem_available_mb") or 0)

    if mem_available_mb and mem_available_mb < 256:
        return 30000
    if load_ratio >= 1:
        return 30000
    if load_ratio >= 0.6:
        return 20000
    return 10000


def load_exam_sidebar_history(
    conn: sqlite3.Connection, exam_filename: str
) -> tuple[list[dict], bool, str]:
    """读取历史榜单，命中缓存时避免重复查库。"""
    current_exam = Path(exam_filename).name
    now = time.time()

    with SIDEBAR_HISTORY_CACHE_LOCK:
        cached = SIDEBAR_HISTORY_CACHE.get(current_exam)
        if cached and now - float(cached.get("updated_at", 0)) < SIDEBAR_HISTORY_TTL_SECONDS:
            return list(cached.get("top_students", [])), True, str(cached.get("updated_iso") or "")

    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT
            st.id as student_id,
            st.class_number,
            st.student_number,
            st.name,
            ranked.best_score
        FROM students st
        JOIN (
            SELECT student_id, MAX(total_score) as best_score
            FROM sessions
            WHERE exam_filename = %s AND status = 'completed'
            GROUP BY student_id
        ) ranked ON ranked.student_id = st.id
        ORDER BY ranked.best_score DESC, st.class_number ASC, st.student_number ASC, st.id ASC
        LIMIT %s
        """,
        (current_exam, SIDEBAR_ACTIVE_LIMIT),
    )

    top_students = []
    for idx, row in enumerate(cursor.fetchall(), start=1):
        item = dict(row)
        item["rank"] = idx
        item["best_score"] = float(item.get("best_score") or 0)
        top_students.append(item)

    updated_iso = datetime.now().isoformat()
    with SIDEBAR_HISTORY_CACHE_LOCK:
        SIDEBAR_HISTORY_CACHE[current_exam] = {
            "updated_at": now,
            "updated_iso": updated_iso,
            "top_students": list(top_students),
        }
    return top_students, False, updated_iso


def load_exam_sidebar_active_students(
    conn: sqlite3.Connection, exam_filename: str
) -> tuple[int, list[dict]]:
    """实时读取当前在线练习学生及其当前得分。"""
    cursor = conn.cursor()
    current_exam = Path(exam_filename).name

    cursor.execute(
        "SELECT COUNT(*) FROM sessions WHERE status = 'active' AND exam_filename = %s",
        (current_exam,),
    )
    active_sessions = _fetchone_value(cursor.fetchone(), 0) or 0

    cursor.execute(
        """
        SELECT
            s.id AS session_id,
            s.student_id,
            s.start_time,
            st.class_number,
            st.student_number,
            st.name,
            COALESCE(SUM(a.score), 0) AS current_score,
            COUNT(a.id) AS answered_count
        FROM sessions s
        JOIN students st ON st.id = s.student_id
        LEFT JOIN answers a ON a.session_id = s.id
        WHERE s.status = 'active' AND s.exam_filename = %s
        GROUP BY s.id, s.student_id, s.start_time, st.class_number, st.student_number, st.name
        ORDER BY current_score DESC, answered_count DESC, s.start_time ASC, s.id ASC
        LIMIT %s
        """,
        (current_exam, SIDEBAR_ACTIVE_LIMIT),
    )

    active_students = []
    for idx, row in enumerate(cursor.fetchall(), start=1):
        item = dict(row)
        item["rank"] = idx
        item["current_score"] = float(item.get("current_score") or 0)
        item["answered_count"] = int(item.get("answered_count") or 0)
        active_students.append(item)

    return int(active_sessions), active_students


def ensure_exam_questions_table(cursor: sqlite3.Cursor) -> None:
    """确保 exam_questions 表存在，用于缓存多份试卷的题目。"""
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS exam_questions (
            exam_filename TEXT NOT NULL,
            id TEXT NOT NULL,
            type TEXT NOT NULL CHECK(type IN ('choice', 'fill')),
            question TEXT NOT NULL,
            score REAL NOT NULL,
            explanation TEXT,
            image TEXT,
            correct_answer TEXT,
            options TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (exam_filename, id)
        )
        """
    )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_exam_questions_exam ON exam_questions(exam_filename)"
    )


def extract_correct_answer(question_data: dict) -> str | int | None:
    """兼容不同试卷格式的答案字段。"""
    question_type = str(question_data.get("type", "")).strip()

    if question_type == "fill":
        fill_blanks = parse_fill_blanks(question_data.get("answers"))
        if not fill_blanks:
            fill_blanks = parse_fill_blanks(question_data.get("correctAnswer"))
        if fill_blanks:
            return json.dumps(fill_blanks, ensure_ascii=False)
        return None

    correct_answer = question_data.get("correctAnswer")
    if correct_answer is None and "answers" in question_data:
        answers = question_data.get("answers", [])
        if answers:
            all_answers = []
            for answer_group in answers:
                if isinstance(answer_group, list):
                    all_answers.extend(
                        [str(a) for a in answer_group if a is not None and a != ""]
                    )
                elif answer_group is not None and answer_group != "":
                    all_answers.append(str(answer_group))
            correct_answer = ",".join(all_answers) if all_answers else None
    return correct_answer


def normalize_image_field(image_value):
    """兼容单图/多图：数据库中 image 可能是字符串或 JSON 数组字符串。"""
    if image_value is None:
        return None
    if isinstance(image_value, list):
        return [str(v) for v in image_value if v]
    if not isinstance(image_value, str):
        return str(image_value)

    raw = image_value.strip()
    if not raw:
        return None

    if raw.startswith("["):
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                return [str(v) for v in parsed if v]
        except Exception:
            pass
    return raw


def _data_uri_extension(header: str) -> str:
    match = re.match(r"data:image/([a-zA-Z0-9.+-]+);base64$", header)
    if not match:
        return "bin"
    subtype = match.group(1).lower()
    if subtype == "jpeg":
        return "jpg"
    if subtype == "svg+xml":
        return "svg"
    return subtype


def externalize_image_value(image_value: Any) -> Any:
    """把 data URI 图片落盘为静态文件，减少题目 JSON 体积。"""
    if isinstance(image_value, list):
        return [externalize_image_value(item) for item in image_value if item]
    if not isinstance(image_value, str):
        return image_value

    raw = image_value.strip()
    if not raw.startswith("data:image/") or ";base64," not in raw:
        return image_value

    header, encoded = raw.split(",", 1)
    digest = hashlib.sha256(raw.encode("utf-8")).hexdigest()
    ext = _data_uri_extension(header)
    GENERATED_IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    image_path = GENERATED_IMAGES_DIR / f"{digest}.{ext}"
    if not image_path.exists():
        try:
            image_path.write_bytes(base64.b64decode(encoded))
        except Exception:
            return image_value
    return f"/xx/generated-images/{image_path.name}"


def externalize_question_images(question: dict) -> dict:
    """把题目主体、选项和填空图统一转换为静态图片 URL。"""
    question["image"] = externalize_image_value(normalize_image_field(question.get("image")))

    if isinstance(question.get("options"), list):
        normalized_options = []
        for option in question["options"]:
            if not isinstance(option, dict):
                normalized_options.append(option)
                continue
            option_copy = dict(option)
            if "image" in option_copy:
                option_copy["image"] = externalize_image_value(option_copy.get("image"))
            if "images" in option_copy:
                option_copy["images"] = externalize_image_value(option_copy.get("images"))
            normalized_options.append(option_copy)
        question["options"] = normalized_options

    if isinstance(question.get("fill_blanks"), list):
        normalized_blanks = []
        for blank in question["fill_blanks"]:
            if not isinstance(blank, dict):
                normalized_blanks.append(blank)
                continue
            blank_copy = dict(blank)
            blank_copy["images"] = externalize_image_value(blank_copy.get("images") or [])
            normalized_blanks.append(blank_copy)
        question["fill_blanks"] = normalized_blanks

    return question


def list_exam_files() -> list[Path]:
    """列出可用试卷文件（根目录 + exams 子目录）。"""
    EXAMS_DIR.mkdir(parents=True, exist_ok=True)
    files_by_name: dict[str, Path] = {}

    for exam_path in sorted(BASE_DIR.glob("*.json")):
        files_by_name[exam_path.name] = exam_path
    for exam_path in sorted(EXAMS_DIR.glob("*.json")):
        files_by_name[exam_path.name] = exam_path

    return [files_by_name[name] for name in sorted(files_by_name.keys())]


def resolve_exam_path(filename: str) -> Path | None:
    """按文件名解析试卷路径。"""
    safe_name = Path(filename).name
    for exam_path in list_exam_files():
        if exam_path.name == safe_name:
            return exam_path
    return None


def load_exam_json(exam_path: Path) -> dict:
    """读取并校验试卷 JSON。"""
    try:
        with open(exam_path, "r", encoding="utf-8") as f:
            data = normalize_exam_json(json.load(f))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"读取试卷失败: {e}") from e

    questions = data.get("questions")
    if not isinstance(questions, list) or not questions:
        raise HTTPException(
            status_code=400, detail="试卷格式错误：questions 必须为非空数组"
        )
    return data


def set_current_exam(conn: sqlite3.Connection, filename: str) -> None:
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO settings (key, value, updated_at)
        VALUES ('current_exam', %s, CURRENT_TIMESTAMP)
        ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = CURRENT_TIMESTAMP
        """,
        (filename,),
    )


def get_current_exam_name(conn: sqlite3.Connection) -> str | None:
    cursor = conn.cursor()
    cursor.execute("SELECT value FROM settings WHERE key = 'current_exam'")
    row = cursor.fetchone()
    if row and row["value"]:
        return row["value"]
    return None


def get_exam_mode(conn: sqlite3.Connection) -> str:
    """获取当前模式：practice 或 exam。默认 practice。"""
    cursor = conn.cursor()
    cursor.execute("SELECT value FROM settings WHERE key = 'exam_mode'")
    row = cursor.fetchone()
    if row and row["value"] in ("practice", "exam"):
        return row["value"]
    return "practice"


def set_exam_mode(conn: sqlite3.Connection, mode: str) -> None:
    if mode not in ("practice", "exam"):
        raise HTTPException(status_code=400, detail="mode 必须为 practice 或 exam")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO settings (key, value, updated_at)
        VALUES ('exam_mode', %s, CURRENT_TIMESTAMP)
        ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = CURRENT_TIMESTAMP
        """,
        (mode,),
    )


def get_exam_distribution(conn: sqlite3.Connection) -> tuple[str, Optional[str]]:
    """
    返回考试分发策略:
    - ('random', None): 随机分发现有试卷
    - ('fixed', filename): 固定分发指定试卷
    """
    cursor = conn.cursor()
    cursor.execute("SELECT value FROM settings WHERE key = 'exam_distribution'")
    dist = cursor.fetchone()
    cursor.execute("SELECT value FROM settings WHERE key = 'exam_fixed'")
    fixed = cursor.fetchone()
    dist_value = dist["value"] if dist and dist["value"] in ("random", "fixed") else "random"
    fixed_value = fixed["value"] if fixed and fixed["value"] else None
    return dist_value, fixed_value


def set_exam_distribution(conn: sqlite3.Connection, mode: str, filename: Optional[str]) -> None:
    if mode not in ("random", "fixed"):
        raise HTTPException(status_code=400, detail="distribution 必须为 random 或 fixed")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO settings (key, value, updated_at)
        VALUES ('exam_distribution', %s, CURRENT_TIMESTAMP)
        ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = CURRENT_TIMESTAMP
        """,
        (mode,),
    )
    cursor.execute(
        """
        INSERT INTO settings (key, value, updated_at)
        VALUES ('exam_fixed', %s, CURRENT_TIMESTAMP)
        ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = CURRENT_TIMESTAMP
        """,
        (filename or ""),
    )


def import_exam_into_questions(conn: sqlite3.Connection, exam_path: Path) -> dict:
    """把选中的试卷导入 questions 表（覆盖当前题库）。"""
    data = load_exam_json(exam_path)
    questions = data["questions"]
    cursor = conn.cursor()

    # 同步到全局 questions（兼容旧前端/教师端）
    cursor.execute("DELETE FROM questions")

    insert_count = 0
    for q in questions:
        if not q.get("id") or not q.get("type") or not q.get("question"):
            raise HTTPException(status_code=400, detail=f"试卷题目字段缺失: {q}")

        options = (
            json.dumps(q.get("options", []), ensure_ascii=False)
            if "options" in q
            else None
        )
        correct_answer = extract_correct_answer(q)
        image_value = q.get("image")
        if not image_value and isinstance(q.get("images"), list):
            image_value = json.dumps(q.get("images", []), ensure_ascii=False)

        cursor.execute(
            """
            INSERT INTO questions (id, type, question, score, explanation, image, correct_answer, options)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                str(q["id"]),
                str(q["type"]),
                str(q["question"]),
                float(q.get("score", 0)),
                q.get("explanation", ""),
                image_value,
                correct_answer,
                options,
            ),
        )
        insert_count += 1

    # 也缓存到 exam_questions（用于多试卷并行）
    import_exam_into_exam_questions(conn, exam_path, data=data)

    set_current_exam(conn, exam_path.name)
    conn.commit()

    return {
        "filename": exam_path.name,
        "title": data.get("title", exam_path.name),
        "question_count": insert_count,
    }


def import_exam_into_exam_questions(
    conn: sqlite3.Connection, exam_path: Path, data: Optional[dict] = None
) -> dict:
    """把试卷题目写入 exam_questions（不影响当前 questions）。"""
    exam_data = data or load_exam_json(exam_path)
    questions = exam_data["questions"]
    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM exam_questions WHERE exam_filename = %s", (exam_path.name,)
    )

    count = 0
    for q in questions:
        if not q.get("id") or not q.get("type") or not q.get("question"):
            raise HTTPException(status_code=400, detail=f"试卷题目字段缺失: {q}")
        options = (
            json.dumps(q.get("options", []), ensure_ascii=False)
            if "options" in q
            else None
        )
        correct_answer = extract_correct_answer(q)
        image_value = q.get("image")
        if not image_value and isinstance(q.get("images"), list):
            image_value = json.dumps(q.get("images", []), ensure_ascii=False)

        cursor.execute(
            """
            INSERT INTO exam_questions (exam_filename, id, type, question, score, explanation, image, correct_answer, options)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                exam_path.name,
                str(q["id"]),
                str(q["type"]),
                str(q["question"]),
                float(q.get("score", 0)),
                q.get("explanation", ""),
                image_value,
                correct_answer,
                options,
            ),
        )
        count += 1

    conn.commit()
    return {
        "filename": exam_path.name,
        "title": exam_data.get("title", exam_path.name),
        "question_count": count,
    }


# 数据模型
class Student(BaseModel):
    id: int
    exam_number: str
    class_number: int
    student_number: int
    name: str
    subject_group: int


class Question(BaseModel):
    id: str
    type: str
    question: str
    score: float
    explanation: Optional[str] = None
    image: Optional[str | List[str]] = None
    correct_answer: Optional[str] = None
    fill_blanks: Optional[List[dict]] = None
    options: Optional[List[dict]] = None


class QuestionOptionInput(BaseModel):
    text: Optional[str] = None
    image: Optional[str] = None


class FillBlankInput(BaseModel):
    answers: List[str] = Field(default_factory=list)
    images: Optional[List[str]] = None
    label: Optional[str] = None


class QuestionUpdateRequest(BaseModel):
    exam_filename: str
    id: str
    type: str
    question: str
    score: float
    explanation: Optional[str] = None
    image: Optional[str] = None
    options: Optional[List[QuestionOptionInput]] = None
    correct_answer: Optional[str | int] = None
    fill_blanks: List[FillBlankInput] = Field(default_factory=list)
    force: bool = False


class AnswerRequest(BaseModel):
    student_id: int
    session_id: int
    question_id: str
    answer: str | int
    answer_time: int  # 答题用时（秒）


class StudyNoteRequest(BaseModel):
    student_id: int
    session_id: int
    question_id: str
    knowledge_points: str
    thinking: str


class StudyNoteLikeRequest(BaseModel):
    student_id: int


class SessionStart(BaseModel):
    student_id: int
    exam_filename: Optional[str] = None
    override_code: Optional[str] = None


class StudentLoginRequest(BaseModel):
    exam_number: str
    name: str


class SessionResponse(BaseModel):
    id: int
    student_id: int
    start_time: str
    status: str
    exam_filename: Optional[str] = None
    resumed: Optional[bool] = None


def question_sort_key(question_id: str) -> tuple:
    question_text = str(question_id or "").strip()
    # 兼容 "1"、"choice-1"、"q12" 等不同题号格式，优先按数字部分排序
    if question_text.isdigit():
        return ("", int(question_text), question_text)

    parts = question_text.split("-", 1)
    prefix = parts[0].strip().lower() if parts else ""
    remainder = parts[1].strip() if len(parts) == 2 else question_text

    match = re.search(r"(\d+)$", remainder)
    if not match:
        match = re.search(r"(\d+)$", question_text)
    num = int(match.group(1)) if match else 0
    return (prefix, num, question_text)


def build_question_payload(q: dict, exam_mode: str) -> Question:
    if q["options"]:
        q["options"] = json.loads(q["options"])
    if exam_mode == "practice":
        if q.get("correct_answer") is not None:
            q["correct_answer"] = str(q["correct_answer"])
        if q.get("type") == "fill":
            q["fill_blanks"] = parse_fill_blanks(q.get("correct_answer"))
    else:
        blanks = parse_fill_blanks(q.get("correct_answer")) if q.get("type") == "fill" else []
        for blank in blanks:
            blank["answers"] = []
        q["correct_answer"] = None
        q["fill_blanks"] = blanks
        q["explanation"] = None
    q = externalize_question_images(q)
    return Question(**q)


def load_question_summaries_for_exam(
    conn: sqlite3.Connection, exam_filename: Optional[str] = None
) -> list[dict]:
    cursor = conn.cursor()
    if exam_filename:
        cursor.execute(
            "SELECT id, type, score FROM exam_questions WHERE exam_filename = %s",
            (Path(exam_filename).name,),
        )
    else:
        cursor.execute("SELECT id, type, score FROM questions ORDER BY id")
    rows = [dict(row) for row in cursor.fetchall()]
    rows.sort(key=lambda row: question_sort_key(row["id"]))
    return rows


def load_question_detail_for_exam(
    conn: sqlite3.Connection,
    exam_mode: str,
    question_id: str,
    exam_filename: Optional[str] = None,
) -> Question:
    cursor = conn.cursor()
    if exam_filename:
        cursor.execute(
            "SELECT * FROM exam_questions WHERE exam_filename = %s AND id = %s",
            (Path(exam_filename).name, question_id),
        )
    else:
        cursor.execute("SELECT * FROM questions WHERE id = %s", (question_id,))
    row = cursor.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="题目不存在")
    return build_question_payload(dict(row), exam_mode)


def load_questions_for_exam(
    conn: sqlite3.Connection,
    exam_mode: str,
    exam_filename: Optional[str] = None,
    question_type: Optional[str] = None,
) -> list[Question]:
    cursor = conn.cursor()
    rows = []
    if exam_filename:
        cursor.execute(
            "SELECT * FROM exam_questions WHERE exam_filename = %s",
            (Path(exam_filename).name,),
        )
        rows = cursor.fetchall()
    else:
        if question_type:
            cursor.execute("SELECT * FROM questions WHERE type = %s", (question_type,))
        else:
            cursor.execute("SELECT * FROM questions ORDER BY id")
        rows = cursor.fetchall()
    normalized_rows = [dict(row) for row in rows]

    order_ids: list[str] = []
    target_exam = Path(exam_filename).name if exam_filename else (get_current_exam_name(conn) or "")
    if target_exam:
        exam_path = resolve_exam_path(target_exam)
        if exam_path:
            try:
                exam_data = load_exam_json(exam_path)
                order_ids = [str(item.get("id")) for item in exam_data.get("questions", []) if item.get("id")]
            except Exception:
                order_ids = []

    if order_ids:
        rank = {qid: idx for idx, qid in enumerate(order_ids)}
        normalized_rows.sort(
            key=lambda row: (rank.get(str(row.get("id")), 10**9), question_sort_key(row["id"]))
        )
    else:
        normalized_rows.sort(key=lambda row: question_sort_key(row["id"]))

    return [build_question_payload(row, exam_mode) for row in normalized_rows]


def build_question_editor_payload(question: Question) -> dict:
    """教师端编辑所需的完整题目结构。"""
    payload = question.dict()
    payload["image"] = normalize_image_field(payload.get("image"))
    if payload.get("type") == "fill":
        payload["correct_answer"] = None
    return payload


def _clean_optional_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_editor_choice_answer(
    value: Any, option_count: int
) -> int:
    normalized = _answer_to_choice_index(value)
    if not isinstance(normalized, int):
        raise HTTPException(status_code=400, detail="选择题答案必须是选项序号")
    if normalized < 0 or normalized >= option_count:
        raise HTTPException(status_code=400, detail="选择题答案超出选项范围")
    return normalized


def _build_question_from_update(data: QuestionUpdateRequest) -> dict:
    question_type = str(data.type or "").strip()
    if question_type not in ("choice", "fill"):
        raise HTTPException(status_code=400, detail="题型必须为 choice 或 fill")

    question_text = str(data.question or "").strip()
    if not question_text:
        raise HTTPException(status_code=400, detail="题干不能为空")

    normalized_question = {
        "id": str(data.id or "").strip(),
        "type": question_type,
        "question": question_text,
        "score": float(data.score or default_score_for_type(question_type)),
        "explanation": str(data.explanation or "").strip(),
    }
    if not normalized_question["id"]:
        raise HTTPException(status_code=400, detail="题目 ID 不能为空")

    image_value = _clean_optional_text(data.image)
    if image_value:
        normalized_question["image"] = image_value

    if question_type == "choice":
        raw_options = data.options or []
        options = []
        for option in raw_options:
            option_text = str(option.text or "").strip()
            option_image = _clean_optional_text(option.image)
            if not option_text and not option_image:
                continue
            options.append({"text": option_text, "image": option_image})
        if not options:
            raise HTTPException(status_code=400, detail="选择题至少需要一个选项")
        normalized_question["options"] = options
        normalized_question["correctAnswer"] = _normalize_editor_choice_answer(
            data.correct_answer, len(options)
        )
        return normalized_question

    raw_blanks = data.fill_blanks or []
    blanks = []
    for blank in raw_blanks:
        answers = _expand_fill_blank_answers(_coerce_str_list(blank.answers))
        images = _coerce_image_list(blank.images or [])
        label = _clean_optional_text(blank.label)
        if answers or images or label:
            blanks.append({"answers": answers, "images": images, "label": label})
    if not blanks:
        raise HTTPException(status_code=400, detail="填空题至少需要一个空和答案")
    normalized_question["answers"] = blanks
    return normalized_question


def write_exam_json(exam_path: Path, data: dict) -> None:
    """把试卷 JSON 原子写回文件。"""
    temp_path = exam_path.with_suffix(exam_path.suffix + ".tmp")
    with open(temp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    temp_path.replace(exam_path)


def ensure_study_note_tables(cursor: Any) -> None:
    """确保复盘笔记与点赞表存在。"""
    if _is_postgres():
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS study_notes (
                id SERIAL PRIMARY KEY,
                student_id INTEGER NOT NULL REFERENCES students(id),
                session_id INTEGER NOT NULL REFERENCES sessions(id),
                question_id TEXT NOT NULL,
                exam_filename TEXT DEFAULT '',
                question_text TEXT NOT NULL,
                knowledge_points TEXT DEFAULT '',
                thinking TEXT DEFAULT '',
                like_count INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(student_id, session_id, question_id)
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS study_note_likes (
                id SERIAL PRIMARY KEY,
                note_id INTEGER NOT NULL REFERENCES study_notes(id) ON DELETE CASCADE,
                student_id INTEGER NOT NULL REFERENCES students(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(note_id, student_id)
            )
            """
        )
    else:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS study_notes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id INTEGER NOT NULL,
                session_id INTEGER NOT NULL,
                question_id TEXT NOT NULL,
                exam_filename TEXT DEFAULT '',
                question_text TEXT NOT NULL,
                knowledge_points TEXT DEFAULT '',
                thinking TEXT DEFAULT '',
                like_count INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(student_id, session_id, question_id)
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS study_note_likes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                note_id INTEGER NOT NULL,
                student_id INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(note_id, student_id)
            )
            """
        )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_study_notes_student ON study_notes(student_id)"
    )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_study_notes_question_text ON study_notes(question_text)"
    )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_study_note_likes_note ON study_note_likes(note_id)"
    )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_study_note_likes_student ON study_note_likes(student_id)"
    )


def serialize_study_note(row: dict, current_student_id: Optional[int] = None) -> dict:
    note = dict(row)
    note["like_count"] = int(note.get("like_count") or 0)
    note["liked_by_me"] = bool(note.get("liked_by_me"))
    note["is_own"] = current_student_id is not None and note.get("student_id") == current_student_id
    return note


# 数据库连接
def get_db():
    if _is_postgres():
        if psycopg is None:
            raise RuntimeError("psycopg 未安装，无法连接 PostgreSQL")
        global PG_POOL
        if PG_POOL is None:
            if ConnectionPool is None:
                raise RuntimeError("psycopg_pool 未安装，无法创建连接池")
            min_size = int(os.getenv("PG_POOL_MIN", "5"))
            max_size = int(os.getenv("PG_POOL_MAX", "60"))
            timeout = float(os.getenv("PG_POOL_TIMEOUT", "10"))
            PG_POOL = ConnectionPool(
                _pg_conninfo(),
                min_size=min_size,
                max_size=max_size,
                timeout=timeout,
                kwargs={"row_factory": dict_row},
            )
        with PG_POOL.connection() as conn:
            yield conn
    else:
        raw_conn = sqlite3.connect(DATABASE, check_same_thread=False, timeout=30)
        raw_conn.row_factory = sqlite3.Row
        raw_conn.execute("PRAGMA busy_timeout = 30000")
        # Keep SQLite settings consistent per-connection under concurrent load
        raw_conn.execute("PRAGMA journal_mode=WAL")
        raw_conn.execute("PRAGMA synchronous=NORMAL")
        raw_conn.execute("PRAGMA temp_store=MEMORY")
        conn = _SQLiteCompatConn(raw_conn)
        try:
            yield conn
        finally:
            raw_conn.close()


# 初始化数据库
@asynccontextmanager
async def lifespan(app: FastAPI):
    # 启动时执行
    init_db()
    _start_answer_worker()
    yield
    # 关闭时执行
    _stop_answer_worker()
    global PG_POOL
    if PG_POOL is not None:
        PG_POOL.close()


app = FastAPI(title="练习系统API", lifespan=lifespan)

# CORS配置
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.api_route(
    "/xx/api/{path:path}",
    methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS", "HEAD"],
)
def local_xx_api_redirect(path: str, request: Request):
    """
    本地直跑 FastAPI 时，前端会请求 /xx/api/*。
    线上由 Nginx 重写到 /api/*；本地没有 Nginx，因此这里做兼容跳转。
    """
    target = f"/api/{path}"
    if request.url.query:
        target = f"{target}?{request.url.query}"
    return RedirectResponse(url=target, status_code=307)


def init_db_postgres():
    """初始化 PostgreSQL 数据库"""
    if psycopg is None:
        raise RuntimeError("psycopg 未安装，无法初始化 PostgreSQL")
    conn = psycopg.connect(_pg_conninfo(), row_factory=dict_row)
    cursor = conn.cursor()
    cursor.execute("SELECT pg_advisory_lock(987654321)")

    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS students (
                id SERIAL PRIMARY KEY,
                exam_number TEXT UNIQUE NOT NULL,
                class_number INTEGER NOT NULL,
                student_number INTEGER NOT NULL,
                name TEXT NOT NULL,
                subject_group INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS questions (
                id TEXT PRIMARY KEY,
                type TEXT NOT NULL CHECK(type IN ('choice', 'fill')),
                question TEXT NOT NULL,
                score REAL NOT NULL,
                explanation TEXT,
                image TEXT,
                correct_answer TEXT,
                options TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS sessions (
                id SERIAL PRIMARY KEY,
                student_id INTEGER NOT NULL REFERENCES students(id),
                start_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                end_time TIMESTAMP,
                total_score REAL DEFAULT 0,
                status TEXT DEFAULT 'active' CHECK(status IN ('active', 'completed', 'abandoned')),
                exam_filename TEXT DEFAULT ''
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS answers (
                id SERIAL PRIMARY KEY,
                session_id INTEGER NOT NULL REFERENCES sessions(id),
                student_id INTEGER NOT NULL REFERENCES students(id),
                question_id TEXT NOT NULL,
                answer TEXT,
                is_correct BOOLEAN,
                score REAL DEFAULT 0,
                answer_time INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                exam_filename TEXT DEFAULT '',
                UNIQUE(session_id, question_id)
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS study_notes (
                id SERIAL PRIMARY KEY,
                student_id INTEGER NOT NULL REFERENCES students(id),
                session_id INTEGER NOT NULL REFERENCES sessions(id),
                question_id TEXT NOT NULL,
                exam_filename TEXT DEFAULT '',
                question_text TEXT NOT NULL,
                knowledge_points TEXT DEFAULT '',
                thinking TEXT DEFAULT '',
                like_count INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(student_id, session_id, question_id)
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS study_note_likes (
                id SERIAL PRIMARY KEY,
                note_id INTEGER NOT NULL REFERENCES study_notes(id) ON DELETE CASCADE,
                student_id INTEGER NOT NULL REFERENCES students(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(note_id, student_id)
            )
            """
        )

        ensure_exam_questions_table(cursor)
        ensure_study_note_tables(cursor)
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_answers_question ON answers(question_id)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_answers_student ON answers(student_id)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_answers_session ON answers(session_id)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_answers_created ON answers(created_at)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_students_exam ON students(exam_number)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_sessions_status_exam ON sessions(status, exam_filename)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_sessions_exam ON sessions(exam_filename)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_answers_exam ON answers(exam_filename)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_study_notes_student ON study_notes(student_id)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_study_notes_question_text ON study_notes(question_text)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_study_note_likes_note ON study_note_likes(note_id)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_study_note_likes_student ON study_note_likes(student_id)"
        )
        cursor.execute(
            "ALTER TABLE answers DROP CONSTRAINT IF EXISTS answers_question_id_fkey"
        )

        cursor.execute(
            "INSERT INTO settings (key, value) VALUES ('exam_mode', 'practice') ON CONFLICT (key) DO NOTHING"
        )
        cursor.execute(
            "INSERT INTO settings (key, value) VALUES ('exam_distribution', 'random') ON CONFLICT (key) DO NOTHING"
        )
        cursor.execute(
            "INSERT INTO settings (key, value) VALUES ('exam_fixed', '') ON CONFLICT (key) DO NOTHING"
        )

        # 启动时保证当前试卷可用
        current_exam = get_current_exam_name(conn)
        available_exams = list_exam_files()
        if current_exam:
            current_exam_path = resolve_exam_path(current_exam)
            if current_exam_path:
                import_exam_into_questions(conn, current_exam_path)
                import_exam_into_exam_questions(conn, current_exam_path)
            elif available_exams:
                import_exam_into_questions(conn, available_exams[0])
                import_exam_into_exam_questions(conn, available_exams[0])
            else:
                set_current_exam(conn, "")
        elif available_exams:
            import_exam_into_questions(conn, available_exams[0])
            import_exam_into_exam_questions(conn, available_exams[0])

        conn.commit()
    finally:
        cursor.execute("SELECT pg_advisory_unlock(987654321)")
        conn.close()


def init_db():
    """初始化数据库"""
    if _is_postgres():
        init_db_postgres()
        return
    conn = sqlite3.connect(DATABASE, check_same_thread=False, timeout=30)
    conn.row_factory = sqlite3.Row
    compat_conn = _SQLiteCompatConn(conn)
    cursor = conn.cursor()
    cursor.execute("PRAGMA journal_mode=WAL")
    cursor.execute("PRAGMA synchronous=NORMAL")

    # 检查表是否已存在
    cursor.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='students'"
    )
    if not cursor.fetchone():
        # 表不存在，执行建表脚本
        with open(BASE_DIR / "database" / "schema.sql", "r", encoding="utf-8") as f:
            conn.executescript(f.read())
        print("数据库初始化完成")
    else:
        print("数据库已存在，跳过初始化")

    # 确保settings表存在
    cursor.execute(
        "CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)"
    )
    cursor.execute(
        "INSERT OR IGNORE INTO settings (key, value) VALUES ('exam_mode', 'practice')"
    )
    cursor.execute(
        "INSERT OR IGNORE INTO settings (key, value) VALUES ('exam_distribution', 'random')"
    )
    cursor.execute(
        "INSERT OR IGNORE INTO settings (key, value) VALUES ('exam_fixed', '')"
    )

    # 历史库结构兼容：补充试卷字段
    ensure_column(cursor, "sessions", "exam_filename", "exam_filename TEXT")
    ensure_column(cursor, "answers", "exam_filename", "exam_filename TEXT")
    ensure_exam_questions_table(cursor)
    ensure_study_note_tables(cursor)
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_sessions_exam ON sessions(exam_filename)"
    )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_answers_exam ON answers(exam_filename)"
    )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_sessions_status_exam ON sessions(status, exam_filename)"
    )
    cursor.execute("UPDATE sessions SET exam_filename = '' WHERE exam_filename IS NULL")
    cursor.execute("UPDATE answers SET exam_filename = '' WHERE exam_filename IS NULL")

    # 启动时保证当前试卷可用
    current_exam = get_current_exam_name(compat_conn)
    available_exams = list_exam_files()
    if current_exam:
        current_exam_path = resolve_exam_path(current_exam)
        if current_exam_path:
            import_exam_into_questions(compat_conn, current_exam_path)
            import_exam_into_exam_questions(compat_conn, current_exam_path)
        elif available_exams:
            import_exam_into_questions(compat_conn, available_exams[0])
            import_exam_into_exam_questions(compat_conn, available_exams[0])
        else:
            set_current_exam(compat_conn, "")
    elif available_exams:
        import_exam_into_questions(compat_conn, available_exams[0])
        import_exam_into_exam_questions(compat_conn, available_exams[0])

    conn.commit()
    conn.close()


# ========== 学生相关接口 ==========


@app.get("/api/students", response_model=List[Student])
def get_students(
    class_number: Optional[int] = None, conn: sqlite3.Connection = Depends(get_db)
):
    """获取学生列表"""
    cursor = conn.cursor()
    if class_number:
        cursor.execute("SELECT * FROM students WHERE class_number = %s", (class_number,))
    else:
        cursor.execute("SELECT * FROM students ORDER BY class_number, student_number")
    rows = cursor.fetchall()
    return [Student(**dict(row)) for row in rows]


@app.get("/api/students/{student_id}", response_model=Student)
def get_student(student_id: int, conn: sqlite3.Connection = Depends(get_db)):
    """获取单个学生信息"""
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM students WHERE id = %s", (student_id,))
    row = cursor.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="学生不存在")
    return Student(**dict(row))


@app.post("/api/students/login")
def student_login(
    data: StudentLoginRequest, conn: sqlite3.Connection = Depends(get_db)
):
    """按准考证号和姓名校验学生身份。"""
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT id, exam_number, class_number, student_number, name, subject_group
        FROM students
        WHERE exam_number = %s AND name = %s
        LIMIT 1
        """,
        (str(data.exam_number).strip(), str(data.name).strip()),
    )
    row = cursor.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="准考证号或姓名错误")
    student = dict(row)
    return {
        "id": student["id"],
        "exam_number": student["exam_number"],
        "class_number": student["class_number"],
        "student_number": student["student_number"],
        "name": student["name"],
    }


# ========== 题目相关接口 ==========


@app.get("/api/questions", response_model=List[Question])
def get_questions(
    type: Optional[str] = None,
    exam_filename: Optional[str] = None,
    conn: sqlite3.Connection = Depends(get_db),
):
    """获取题目列表"""
    try:
        exam_mode = get_exam_mode(conn)
        return load_questions_for_exam(conn, exam_mode, exam_filename, type)
    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/questions/{question_id}", response_model=Question)
def get_question(
    question_id: str,
    exam_filename: Optional[str] = None,
    conn: sqlite3.Connection = Depends(get_db),
):
    """获取单道题目"""
    exam_mode = get_exam_mode(conn)
    return load_question_detail_for_exam(conn, exam_mode, question_id, exam_filename)


@app.get("/api/exams/{filename}/questions/{question_id}")
def get_exam_question_for_edit(
    filename: str, question_id: str, conn: sqlite3.Connection = Depends(get_db)
):
    """教师端读取指定试卷的完整题目内容。"""
    exam_path = resolve_exam_path(filename)
    if not exam_path:
        raise HTTPException(status_code=404, detail="试卷文件不存在")
    question = load_question_detail_for_exam(conn, "practice", question_id, exam_path.name)
    return build_question_editor_payload(question)


@app.put("/api/exams/{filename}/questions/{question_id}")
def update_exam_question(
    filename: str,
    question_id: str,
    data: QuestionUpdateRequest,
    conn: sqlite3.Connection = Depends(get_db),
):
    """教师端修改试卷题目，并写回服务器 JSON 文件。"""
    exam_path = resolve_exam_path(filename)
    if not exam_path:
        raise HTTPException(status_code=404, detail="试卷文件不存在")

    safe_filename = exam_path.name
    if Path(data.exam_filename).name != safe_filename:
        raise HTTPException(status_code=400, detail="请求中的试卷文件名不匹配")
    if str(data.id or "").strip() != str(question_id).strip():
        raise HTTPException(status_code=400, detail="请求中的题目 ID 不匹配")

    current_exam = get_current_exam_name(conn) or ""
    cursor = conn.cursor()
    if current_exam == safe_filename and not data.force:
        cursor.execute(
            "SELECT COUNT(*) FROM sessions WHERE status = 'active' AND exam_filename = %s",
            (safe_filename,),
        )
        active_count = int(_fetchone_value(cursor.fetchone(), 0) or 0)
        if active_count > 0:
            raise HTTPException(
                status_code=409,
                detail=f"当前有 {active_count} 个正在答题的会话，确认后请勾选强制保存",
            )

    exam_data = load_exam_json(exam_path)
    questions = exam_data.get("questions", [])
    target_index = -1
    for index, question in enumerate(questions):
        if str(question.get("id")) == str(question_id):
            target_index = index
            break
    if target_index < 0:
        raise HTTPException(status_code=404, detail="题目不存在")

    normalized_question = _build_question_from_update(data)
    questions[target_index] = normalized_question
    exam_data["questions"] = questions
    if "title" not in exam_data or not str(exam_data.get("title") or "").strip():
        exam_data["title"] = exam_path.stem

    write_exam_json(exam_path, exam_data)

    import_exam_into_exam_questions(conn, exam_path, data=exam_data)
    synced_to_current = False
    if current_exam == safe_filename:
        import_exam_into_questions(conn, exam_path)
        synced_to_current = True
    else:
        conn.commit()

    refreshed = load_question_detail_for_exam(conn, "practice", question_id, safe_filename)
    return {
        "message": "题目已保存到服务器",
        "filename": safe_filename,
        "question": build_question_editor_payload(refreshed),
        "synced_to_current": synced_to_current,
    }


# ========== 练习会话接口 ==========


@app.post("/api/sessions/start")
def start_session(data: SessionStart, conn: sqlite3.Connection = Depends(get_db)):
    """开始练习会话"""
    cursor = conn.cursor()
    exam_mode = get_exam_mode(conn)

    def build_start_payload(
        session_id: int,
        student_id: int,
        start_time: Any,
        status: str,
        exam_filename: str,
        resumed: bool,
    ) -> dict:
        question_summaries = load_question_summaries_for_exam(conn, exam_filename)
        answers = []
        total_score = 0.0
        if resumed:
            cursor.execute(
                """
                SELECT question_id, answer, is_correct, score
                FROM answers
                WHERE session_id = %s
                ORDER BY created_at ASC
                """,
                (session_id,),
            )
            answers = [dict(row) for row in cursor.fetchall()]
            total_score = sum(float(a.get("score") or 0) for a in answers)
        answered_question_ids = {str(a.get("question_id")) for a in answers}
        initial_question_index = 0
        if question_summaries:
            for idx, summary in enumerate(question_summaries):
                if summary["id"] not in answered_question_ids:
                    initial_question_index = idx
                    break
        return {
            "id": session_id,
            "student_id": student_id,
            "start_time": str(start_time),
            "status": status,
            "exam_filename": exam_filename,
            "resumed": resumed,
            "question_summaries": question_summaries,
            "initial_question_index": initial_question_index,
            "resume_state": {
                "session_id": session_id,
                "exam_filename": exam_filename,
                "answers": answers,
                "total_score": total_score,
            },
        }

    # 如果已有进行中的会话，优先恢复（除非提供口令强制重开）
    cursor.execute(
        """
        SELECT id, student_id, start_time, status, exam_filename
        FROM sessions
        WHERE student_id = %s AND status = 'active'
        ORDER BY start_time DESC
        LIMIT 1
        """,
        (data.student_id,),
    )
    existing = cursor.fetchone()
    if existing and data.override_code != "2055350":
        existing_exam = (existing["exam_filename"] or "").strip()
        if not existing_exam:
            existing_exam = get_current_exam_name(conn) or ""
            cursor.execute(
                "UPDATE sessions SET exam_filename = %s WHERE id = %s",
                (existing_exam, existing["id"]),
            )
            conn.commit()
        if existing_exam:
            cursor.execute(
                "SELECT COUNT(*) FROM exam_questions WHERE exam_filename = %s",
                (existing_exam,),
            )
        if (_fetchone_value(cursor.fetchone(), 0) or 0) == 0:
                path = resolve_exam_path(existing_exam)
                if path:
                    import_exam_into_exam_questions(conn, path)
        return build_start_payload(
            session_id=existing["id"],
            student_id=existing["student_id"],
            start_time=existing["start_time"],
            status=existing["status"],
            exam_filename=existing_exam,
            resumed=True,
        )

    # 确定本次会话使用的试卷
    available_exams = list_exam_files()
    if not available_exams:
        raise HTTPException(status_code=400, detail="当前无可用试卷，请先在教师端导入")

    if exam_mode == "practice":
        # 学生自选试卷（未传则使用当前/首个）
        chosen = data.exam_filename or get_current_exam_name(conn)
        if not chosen and available_exams:
            chosen = available_exams[0].name
        exam_filename = Path(chosen).name if chosen else None
    else:
        # 考试模式：按策略分发
        dist, fixed = get_exam_distribution(conn)
        if dist == "fixed" and fixed:
            exam_filename = fixed
        else:
            # 随机分发现有试卷
            import random

            exam_filename = random.choice(available_exams).name

    if not exam_filename or not resolve_exam_path(exam_filename):
        raise HTTPException(status_code=404, detail="试卷文件不存在")

    # 所有模式：1小时内仅允许一次会话，除非提供口令
    one_hour_ago = datetime.now() - timedelta(hours=1)
    cursor.execute(
        """
        SELECT COUNT(*) FROM sessions
        WHERE student_id = %s AND start_time >= %s AND status IN ('active', 'completed')
        """,
        (data.student_id, one_hour_ago),
    )
    recent_count = _fetchone_value(cursor.fetchone(), 0) or 0
    if recent_count > 0 and data.override_code != "2055350":
        raise HTTPException(
            status_code=429,
            detail="每小时仅允许登录一次。如需重登请联系教师获取口令。",
        )

    # 确保 exam_questions 有缓存，没有则导入
    cursor.execute(
        "SELECT COUNT(*) FROM exam_questions WHERE exam_filename = %s",
        (exam_filename,),
    )
    if (_fetchone_value(cursor.fetchone(), 0) or 0) == 0:
        path = resolve_exam_path(exam_filename)
        if not path:
            raise HTTPException(status_code=404, detail="试卷文件不存在")
        import_exam_into_exam_questions(conn, path)

    # 检查是否有进行中的会话（口令强制重开时才会走到这里）
    cursor.execute(
        "SELECT id FROM sessions WHERE student_id = %s AND status = 'active'",
        (data.student_id,),
    )
    existing = cursor.fetchone()
    if existing:
        cursor.execute(
            "UPDATE sessions SET status = 'abandoned', end_time = %s WHERE id = %s",
            (datetime.now(), existing["id"]),
        )

    # 创建新会话
    if _is_postgres():
        cursor.execute(
            "INSERT INTO sessions (student_id, start_time, status, exam_filename) VALUES (%s, %s, 'active', %s) RETURNING id",
            (data.student_id, datetime.now(), exam_filename),
        )
        session_id = _fetchone_value(cursor.fetchone())
    else:
        cursor.execute(
            "INSERT INTO sessions (student_id, start_time, status, exam_filename) VALUES (%s, %s, 'active', %s)",
            (data.student_id, datetime.now(), exam_filename),
        )
        session_id = cursor.lastrowid
    conn.commit()
    return build_start_payload(
        session_id=session_id,
        student_id=data.student_id,
        start_time=datetime.now().isoformat(),
        status="active",
        exam_filename=exam_filename or "",
        resumed=False,
    )


@app.get("/api/sessions/{session_id}/state")
def get_session_state(
    session_id: int, conn: sqlite3.Connection = Depends(get_db)
):
    """获取会话已答题目与得分，用于恢复进度。"""
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, student_id, status, exam_filename FROM sessions WHERE id = %s",
        (session_id,),
    )
    session = cursor.fetchone()
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")
    if session["status"] != "active":
        raise HTTPException(status_code=400, detail="会话已结束，不能恢复")

    cursor.execute(
        """
        SELECT question_id, answer, is_correct, score
        FROM answers
        WHERE session_id = %s
        ORDER BY created_at ASC
        """,
        (session_id,),
    )
    answers = [dict(row) for row in cursor.fetchall()]
    total_score = sum(float(a.get("score") or 0) for a in answers)

    return {
        "session_id": session_id,
        "exam_filename": (session["exam_filename"] or "").strip(),
        "answers": answers,
        "total_score": total_score,
    }


@app.post("/api/sessions/{session_id}/end")
def end_session(session_id: int, conn: sqlite3.Connection = Depends(get_db)):
    """结束练习会话"""
    cursor = conn.cursor()

    # 计算总分
    cursor.execute("SELECT SUM(score) FROM answers WHERE session_id = %s", (session_id,))
    total_score = _fetchone_value(cursor.fetchone(), 0) or 0

    cursor.execute(
        "UPDATE sessions SET end_time = %s, total_score = %s, status = 'completed' WHERE id = %s",
        (datetime.now(), total_score, session_id),
    )
    conn.commit()

    return {"message": "练习已结束", "total_score": total_score}


# ========== 答题接口 ==========


@app.post("/api/answers/submit")
def submit_answer(data: AnswerRequest, conn: sqlite3.Connection = Depends(get_db)):
    """提交答案"""
    cursor = conn.cursor()
    cursor.execute(
        "SELECT status, exam_filename FROM sessions WHERE id = %s AND student_id = %s",
        (data.session_id, data.student_id),
    )
    session = cursor.fetchone()
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")
    if session["status"] != "active":
        raise HTTPException(status_code=400, detail="会话已结束，不能继续答题")

    session_exam = session["exam_filename"] or get_current_exam_name(conn)

    # 获取题目正确答案（优先 exam_questions）
    cursor.execute(
        "SELECT correct_answer, score, type FROM exam_questions WHERE exam_filename = %s AND id = %s",
        (session_exam or "", data.question_id),
    )
    question = cursor.fetchone()
    if not question:
        cursor.execute(
            "SELECT correct_answer, score, type FROM questions WHERE id = %s",
            (data.question_id,),
        )
        question = cursor.fetchone()
    if not question:
        raise HTTPException(status_code=404, detail="题目不存在")

    # 判断答案是否正确
    # 填空题：每个空单独评分
    if question["type"] == "fill":
        fill_blanks = parse_fill_blanks(question["correct_answer"])
        blank_count = len(fill_blanks)
        student_answers = str(data.answer).split("|||")
        total_similarity = 0.0

        if blank_count > 0:
            for i, blank in enumerate(fill_blanks):
                student_answer = student_answers[i] if i < len(student_answers) else ""
                max_similarity = 0.0
                for raw_answer in blank.get("answers", []):
                    similarity = is_fill_answer_match(student_answer, raw_answer)
                    if similarity > max_similarity:
                        max_similarity = similarity
                total_similarity += max_similarity

            score = round(question["score"] * total_similarity / blank_count, 2)
            is_correct = (
                total_similarity >= blank_count - 0.0001
            )  # Allow small floating point error
        else:
            score = 0
            is_correct = False
    else:
        # 选择题：整体判定
        correct_answers = str(question["correct_answer"]).replace("，", ",").split(",")
        student_answer = str(data.answer).strip()
        is_correct = any(
            correct.strip() == student_answer for correct in correct_answers
        )
        score = question["score"] if is_correct else 0

    # 保存或更新答案（PostgreSQL 使用异步批量写入降低写入压力）
    record = (
        data.session_id,
        data.student_id,
        data.question_id,
        data.answer,
        is_correct,
        score,
        data.answer_time,
        datetime.now(),
        session_exam or "",
    )
    queued = False
    if _is_postgres() and ANSWER_QUEUE is not None:
        try:
            ANSWER_QUEUE.put_nowait(record)
            queued = True
        except queue.Full:
            queued = False

    if not queued:
        cursor.execute(
            """
            INSERT INTO answers
            (session_id, student_id, question_id, answer, is_correct, score, answer_time, created_at, exam_filename)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT(session_id, question_id) DO UPDATE SET
                answer = excluded.answer,
                is_correct = excluded.is_correct,
                score = excluded.score,
                answer_time = excluded.answer_time,
                created_at = excluded.created_at,
                exam_filename = excluded.exam_filename
            """,
            record,
        )
        conn.commit()

    if get_exam_mode(conn) == "exam":
        return {"score": score}

    # Calculate correct_count as the number of blanks with >= 90% similarity (full score)
    correct_count_value = None
    blank_count_value = None
    if question["type"] == "fill":
        correct_count_value = 0
        if blank_count > 0:
            for i, blank in enumerate(fill_blanks):
                student_answer = student_answers[i] if i < len(student_answers) else ""
                max_similarity = 0.0
                for raw_answer in blank.get("answers", []):
                    similarity = is_fill_answer_match(student_answer, raw_answer)
                    if similarity > max_similarity:
                        max_similarity = similarity
                if max_similarity >= 0.9:
                    correct_count_value += 1
        blank_count_value = blank_count

    return {
        "is_correct": is_correct,
        "score": score,
        "correct_answer": question["correct_answer"],
        "correct_count": correct_count_value,
        "blank_count": blank_count_value,
    }


@app.get("/api/notes")
def get_study_notes(
    question_text: str,
    current_student_id: Optional[int] = None,
    limit: int = 20,
    conn: sqlite3.Connection = Depends(get_db),
):
    """获取同题复盘笔记（跨试卷可见）。"""
    cursor = conn.cursor()
    safe_limit = min(max(limit, 1), 50)
    cursor.execute(
        """
        SELECT
            n.id,
            n.student_id,
            n.session_id,
            n.question_id,
            n.exam_filename,
            n.question_text,
            n.knowledge_points,
            n.thinking,
            n.like_count,
            n.created_at,
            n.updated_at,
            s.name,
            s.class_number,
            s.student_number,
            CASE
                WHEN %s IS NULL THEN 0
                WHEN EXISTS (
                    SELECT 1
                    FROM study_note_likes l
                    WHERE l.note_id = n.id AND l.student_id = %s
                ) THEN 1
                ELSE 0
            END AS liked_by_me
        FROM study_notes n
        JOIN students s ON s.id = n.student_id
        WHERE n.question_text = %s
        ORDER BY n.like_count DESC, n.updated_at DESC, n.created_at DESC
        LIMIT %s
        """,
        (current_student_id, current_student_id, question_text, safe_limit),
    )
    rows = [serialize_study_note(dict(row), current_student_id) for row in cursor.fetchall()]
    return {"notes": rows}


@app.post("/api/notes")
def save_study_note(
    data: StudyNoteRequest, conn: sqlite3.Connection = Depends(get_db)
):
    """保存学生对已答题目的知识点与思路复盘。"""
    if get_exam_mode(conn) != "practice":
        raise HTTPException(status_code=400, detail="当前仅练习模式允许提交复盘")

    knowledge_points = str(data.knowledge_points or "").strip()
    thinking = str(data.thinking or "").strip()
    if not knowledge_points and not thinking:
        raise HTTPException(status_code=400, detail="请至少填写知识点或答题思路")

    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, student_id, exam_filename FROM sessions WHERE id = %s",
        (data.session_id,),
    )
    session = cursor.fetchone()
    if not session or session["student_id"] != data.student_id:
        raise HTTPException(status_code=404, detail="会话不存在")

    cursor.execute(
        """
        SELECT 1
        FROM answers
        WHERE session_id = %s AND student_id = %s AND question_id = %s
        """,
        (data.session_id, data.student_id, data.question_id),
    )
    if not cursor.fetchone():
        raise HTTPException(status_code=400, detail="请先提交本题答案，再写复盘")

    session_exam = session["exam_filename"] or ""
    cursor.execute(
        "SELECT question FROM exam_questions WHERE exam_filename = %s AND id = %s",
        (session_exam, data.question_id),
    )
    question = cursor.fetchone()
    if not question:
        cursor.execute("SELECT question FROM questions WHERE id = %s", (data.question_id,))
        question = cursor.fetchone()
    if not question:
        raise HTTPException(status_code=404, detail="题目不存在")

    question_text = str(question["question"] or "").strip()
    now = datetime.now()
    cursor.execute(
        """
        INSERT INTO study_notes
        (student_id, session_id, question_id, exam_filename, question_text, knowledge_points, thinking, created_at, updated_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT(student_id, session_id, question_id) DO UPDATE SET
            knowledge_points = excluded.knowledge_points,
            thinking = excluded.thinking,
            exam_filename = excluded.exam_filename,
            question_text = excluded.question_text,
            updated_at = excluded.updated_at
        """,
        (
            data.student_id,
            data.session_id,
            data.question_id,
            session_exam,
            question_text,
            knowledge_points,
            thinking,
            now,
            now,
        ),
    )
    conn.commit()

    cursor.execute(
        """
        SELECT
            n.id,
            n.student_id,
            n.session_id,
            n.question_id,
            n.exam_filename,
            n.question_text,
            n.knowledge_points,
            n.thinking,
            n.like_count,
            n.created_at,
            n.updated_at,
            s.name,
            s.class_number,
            s.student_number,
            0 AS liked_by_me
        FROM study_notes n
        JOIN students s ON s.id = n.student_id
        WHERE n.student_id = %s AND n.session_id = %s AND n.question_id = %s
        """,
        (data.student_id, data.session_id, data.question_id),
    )
    row = cursor.fetchone()
    return {"note": serialize_study_note(dict(row), data.student_id)}


@app.post("/api/notes/{note_id}/like")
def like_study_note(
    note_id: int,
    data: StudyNoteLikeRequest,
    conn: sqlite3.Connection = Depends(get_db),
):
    """给复盘笔记点赞；同一学生对同一条笔记只记一次。"""
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, student_id, like_count FROM study_notes WHERE id = %s",
        (note_id,),
    )
    note = cursor.fetchone()
    if not note:
        raise HTTPException(status_code=404, detail="复盘不存在")
    if note["student_id"] == data.student_id:
        raise HTTPException(status_code=400, detail="不能给自己的复盘点赞")

    cursor.execute(
        """
        INSERT INTO study_note_likes (note_id, student_id, created_at)
        VALUES (%s, %s, %s)
        ON CONFLICT(note_id, student_id) DO NOTHING
        """,
        (note_id, data.student_id, datetime.now()),
    )
    inserted = getattr(cursor, "rowcount", 0) or 0
    if inserted > 0:
        cursor.execute(
            "UPDATE study_notes SET like_count = like_count + 1, updated_at = %s WHERE id = %s",
            (datetime.now(), note_id),
        )
    conn.commit()

    cursor.execute("SELECT like_count FROM study_notes WHERE id = %s", (note_id,))
    like_count = int(_fetchone_value(cursor.fetchone(), 0) or 0)
    return {"note_id": note_id, "like_count": like_count, "liked": inserted > 0}


# ========== 试卷管理接口 ==========


@app.get("/api/exams/list")
def get_exam_list(conn: sqlite3.Connection = Depends(get_db)):
    """获取可用的试卷列表"""
    exams = []
    for exam_path in list_exam_files():
        try:
            data = load_exam_json(exam_path)
            exams.append(
                {
                    "filename": exam_path.name,
                    "title": data.get("title", exam_path.name),
                    "question_count": len(data.get("questions", [])),
                }
            )
        except Exception:
            pass

    return exams


@app.get("/api/exams/current")
def get_current_exam(conn: sqlite3.Connection = Depends(get_db)):
    """获取当前使用的试卷"""
    current_exam = get_current_exam_name(conn)
    return {"current_exam": current_exam}


class ExamSwitch(BaseModel):
    filename: str


class ExamModeSwitch(BaseModel):
    mode: str


class ExamDistributionSwitch(BaseModel):
    distribution: str  # random | fixed
    filename: Optional[str] = None


@app.post("/api/exams/switch")
def switch_exam(data: ExamSwitch, conn: sqlite3.Connection = Depends(get_db)):
    """切换当前使用的试卷"""
    filename = Path(data.filename).name
    exam_path = resolve_exam_path(filename)
    if not exam_path:
        raise HTTPException(status_code=404, detail="试卷文件不存在")

    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM sessions WHERE status = 'active'")
    active_count = _fetchone_value(cursor.fetchone(), 0) or 0
    if active_count > 0:
        cursor.execute(
            "UPDATE sessions SET status = 'abandoned', end_time = %s WHERE status = 'active'",
            (datetime.now(),),
        )

    import_result = import_exam_into_questions(conn, exam_path)
    return {
        "message": "试卷切换成功",
        "current_exam": import_result["filename"],
        "question_count": import_result["question_count"],
        "abandoned_sessions": active_count,
    }


@app.get("/api/exams/mode")
def get_exam_mode_api(conn: sqlite3.Connection = Depends(get_db)):
    return {"mode": get_exam_mode(conn)}


@app.post("/api/exams/mode")
def set_exam_mode_api(data: ExamModeSwitch, conn: sqlite3.Connection = Depends(get_db)):
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM sessions WHERE status = 'active'")
    active_count = _fetchone_value(cursor.fetchone(), 0) or 0
    if active_count > 0:
        cursor.execute(
            "UPDATE sessions SET status = 'abandoned', end_time = %s WHERE status = 'active'",
            (datetime.now(),),
        )
        conn.commit()

    set_exam_mode(conn, data.mode)
    return {"mode": data.mode, "abandoned_sessions": active_count}


@app.get("/api/exams/distribution")
def get_exam_distribution_api(conn: sqlite3.Connection = Depends(get_db)):
    dist, fixed = get_exam_distribution(conn)
    return {"distribution": dist, "filename": fixed}


@app.post("/api/exams/distribution")
def set_exam_distribution_api(
    data: ExamDistributionSwitch, conn: sqlite3.Connection = Depends(get_db)
):
    dist = data.distribution
    filename = data.filename
    if dist == "fixed":
        if not filename:
            raise HTTPException(status_code=400, detail="固定分发必须指定试卷文件名")
        if not resolve_exam_path(filename):
            raise HTTPException(status_code=404, detail="指定试卷不存在")

    set_exam_distribution(conn, dist, filename)
    return {"distribution": dist, "filename": filename or ""}


@app.get("/api/health")
def get_health_status():
    """系统轻量健康检查，用于前端提示教师处理。"""
    return get_system_health()


@app.post("/api/exams/upload")
async def upload_exam(
    file: UploadFile = File(...), conn: sqlite3.Connection = Depends(get_db)
):
    """上传新的试卷 JSON 文件。"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")

    safe_name = Path(file.filename).name
    if not safe_name.lower().endswith(".json"):
        raise HTTPException(status_code=400, detail="仅支持 JSON 试卷文件")

    EXAMS_DIR.mkdir(parents=True, exist_ok=True)
    target_path = EXAMS_DIR / safe_name
    existed = target_path.exists()

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="文件内容为空")

    try:
        data = normalize_exam_json(json.loads(content.decode("utf-8")))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"JSON 解析失败: {e}") from e

    questions = data.get("questions")
    if not isinstance(questions, list) or not questions:
        raise HTTPException(
            status_code=400, detail="试卷格式错误：questions 必须为非空数组"
        )

    with open(target_path, "wb") as f:
        f.write(content)

    # 缓存到 exam_questions，便于多试卷并行练习/考试
    import_exam_into_exam_questions(conn, target_path, data=data)

    current_exam = get_current_exam_name(conn)
    abandoned_sessions = 0
    synced_to_db = False
    if current_exam == safe_name:
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM sessions WHERE status = 'active'")
        abandoned_sessions = _fetchone_value(cursor.fetchone(), 0) or 0
        if abandoned_sessions > 0:
            cursor.execute(
                "UPDATE sessions SET status = 'abandoned', end_time = %s WHERE status = 'active'",
                (datetime.now(),),
            )
            conn.commit()

        # 当前试卷被同名覆盖后，立即重导入题库，保证数据库与文件一致
        import_exam_into_questions(conn, target_path)
        synced_to_db = True

    if existed and synced_to_db:
        message = "同名试卷已覆盖，且已同步更新当前考试题库"
    elif existed:
        message = "同名试卷已覆盖"
    else:
        message = "试卷上传成功"

    return {
        "message": message,
        "filename": target_path.name,
        "title": data.get("title", target_path.name),
        "question_count": len(questions),
        "overwritten": existed,
        "synced_to_db": synced_to_db,
        "abandoned_sessions": abandoned_sessions,
    }


# ========== 统计分析接口（核心功能） ==========


@app.get("/api/analysis/question/{question_id}")
def get_question_analysis(question_id: str, conn: sqlite3.Connection = Depends(get_db)):
    """
    获取某道题目的答题情况分析
    包括：哪些学生答了、答案是什么、是否正确、用时等
    """
    cursor = conn.cursor()
    current_exam = get_current_exam_name(conn) or ""

    # 获取题目信息
    cursor.execute("SELECT * FROM questions WHERE id = %s", (question_id,))
    question = cursor.fetchone()
    if not question:
        raise HTTPException(status_code=404, detail="题目不存在")

    # 获取所有答题记录
    cursor.execute(
        """
        SELECT 
            s.name,
            s.class_number,
            s.student_number,
            a.answer,
            a.is_correct,
            a.score,
            a.answer_time,
            a.created_at
        FROM answers a
        JOIN students s ON a.student_id = s.id
        WHERE a.question_id = %s AND a.exam_filename = %s
        ORDER BY a.created_at DESC
        """,
        (question_id, current_exam),
    )
    answers = [dict(row) for row in cursor.fetchall()]

    # 统计信息
    total = len(answers)
    correct = sum(1 for a in answers if a["is_correct"])
    wrong = total - correct
    correct_rate = round(100.0 * correct / total, 2) if total > 0 else 0

    # 各选项选择人数统计（选择题）
    option_stats = {}
    if question["type"] == "choice":
        for ans in answers:
            opt = ans["answer"]
            option_stats[opt] = option_stats.get(opt, 0) + 1

    return {
        "question": {
            "id": question["id"],
            "type": question["type"],
            "content": question["question"],
            "correct_answer": question["correct_answer"],
        },
        "statistics": {
            "total_attempts": total,
            "correct_count": correct,
            "wrong_count": wrong,
            "correct_rate": correct_rate,
            "option_distribution": option_stats,
        },
        "answers": answers,
    }


@app.get("/api/analysis/student/{student_id}")
def get_student_analysis(
    student_id: int,
    include_attempts: bool = True,
    conn: sqlite3.Connection = Depends(get_db),
):
    """获取某个学生的答题分析"""
    cursor = conn.cursor()
    current_exam = get_current_exam_name(conn) or ""

    # 获取学生信息
    cursor.execute("SELECT * FROM students WHERE id = %s", (student_id,))
    student = cursor.fetchone()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    # 获取答题统计
    cursor.execute(
        """
        SELECT 
            COUNT(*) as total_answered,
            SUM(CASE WHEN is_correct IS TRUE THEN 1 ELSE 0 END) as correct_count,
            SUM(score) as total_score,
            AVG(answer_time) as avg_time,
            (
                SELECT COUNT(*)
                FROM study_notes sn
                WHERE sn.student_id = %s
            ) as note_count,
            (
                SELECT COALESCE(SUM(sn.like_count), 0)
                FROM study_notes sn
                WHERE sn.student_id = %s
            ) as likes_received
        FROM answers
        WHERE student_id = %s AND exam_filename = %s
        """,
        (student_id, student_id, student_id, current_exam),
    )
    stats = dict(cursor.fetchone())

    details = []
    attempts = []

    if include_attempts:
        # 读取该学生该试卷下的所有考试会话（每次考试一条记录）
        cursor.execute(
            """
            SELECT id, start_time, end_time, total_score, status, exam_filename
            FROM sessions
            WHERE student_id = %s AND exam_filename = %s
            ORDER BY start_time ASC
            """,
            (student_id, current_exam),
        )
        session_rows = [dict(row) for row in cursor.fetchall()]

        for idx, session_row in enumerate(session_rows, start=1):
            cursor.execute(
                """
                SELECT 
                    q.id,
                    q.question,
                    q.type,
                    q.correct_answer,
                    a.answer,
                    a.is_correct,
                    a.score,
                    a.answer_time,
                    a.created_at
                FROM answers a
                JOIN questions q ON a.question_id = q.id
                WHERE a.session_id = %s AND a.exam_filename = %s
                ORDER BY a.created_at ASC
                """,
                (session_row["id"], current_exam),
            )
            answer_rows = [dict(row) for row in cursor.fetchall()]

            answered_count = len(answer_rows)
            correct_count = sum(1 for a in answer_rows if a.get("is_correct"))
            avg_time = (
                sum(a.get("answer_time") or 0 for a in answer_rows) / answered_count
                if answered_count > 0
                else 0
            )

            attempts.append(
                {
                    "session_id": session_row["id"],
                    "attempt_no": idx,
                    "exam_filename": session_row.get("exam_filename") or current_exam,
                    "start_time": session_row.get("start_time"),
                    "end_time": session_row.get("end_time"),
                    "status": session_row.get("status"),
                    "total_score": session_row.get("total_score") or 0,
                    "answered_count": answered_count,
                    "correct_count": correct_count,
                    "avg_time": avg_time,
                    "answers": answer_rows,
                }
            )

        # 最新尝试排在前面
        attempts = list(reversed(attempts))

    # 兼容旧前端：保留当前试卷下的答题明细（不分次）
    cursor.execute(
        """
        SELECT 
            q.id,
            q.question,
            q.type,
            q.correct_answer,
            a.answer,
            a.is_correct,
            a.score,
            a.answer_time,
            a.created_at
        FROM answers a
        JOIN questions q ON a.question_id = q.id
        WHERE a.student_id = %s AND a.exam_filename = %s
        ORDER BY a.created_at DESC
        """,
        (student_id, current_exam),
    )
    details = [dict(row) for row in cursor.fetchall()]

    return {
        "student": {
            "id": student["id"],
            "name": student["name"],
            "class": student["class_number"],
            "exam_number": student["exam_number"],
        },
        "statistics": stats,
        "attempts": attempts,
        "answers": details,
    }


@app.get("/api/analysis/overview")
def get_overview(
    exam_filename: Optional[str] = None, conn: sqlite3.Connection = Depends(get_db)
):
    """获取整体统计概览"""
    cursor = conn.cursor()
    current_exam = Path(exam_filename).name if exam_filename else get_current_exam_name(conn) or ""

    # 总参与人数
    cursor.execute(
        "SELECT COUNT(DISTINCT student_id) FROM answers WHERE exam_filename = %s",
        (current_exam,),
    )
    total_students = _fetchone_value(cursor.fetchone(), 0)

    # 总答题次数
    cursor.execute(
        "SELECT COUNT(*) FROM answers WHERE exam_filename = %s",
        (current_exam,),
    )
    total_answers = _fetchone_value(cursor.fetchone(), 0)

    # 正在答题人数
    cursor.execute(
        "SELECT COUNT(*) FROM sessions WHERE status = 'active' AND exam_filename = %s",
        (current_exam,),
    )
    active_sessions = _fetchone_value(cursor.fetchone(), 0) or 0

    # 每道题的统计
    cursor.execute(
        """
        SELECT 
            q.id,
            q.question,
            COUNT(a.id) as attempt_count,
            SUM(CASE WHEN a.is_correct IS TRUE THEN 1 ELSE 0 END) as correct_count,
            CASE
                WHEN COUNT(a.id) = 0 THEN 0
                ELSE ROUND(100.0 * SUM(CASE WHEN a.is_correct IS TRUE THEN 1 ELSE 0 END) / COUNT(a.id), 2)
            END as correct_rate
        FROM questions q
        LEFT JOIN answers a ON q.id = a.question_id AND a.exam_filename = %s
        GROUP BY q.id
        ORDER BY correct_rate ASC
        """,
        (current_exam,),
    )
    question_stats = [dict(row) for row in cursor.fetchall()]

    return {
        "total_students": total_students,
        "total_answers": total_answers,
        "active_sessions": active_sessions,
        "question_statistics": question_stats,
    }


@app.get("/api/analysis/active_sessions")
def get_active_sessions(
    exam_filename: Optional[str] = None, conn: sqlite3.Connection = Depends(get_db)
):
    """获取正在答题的学生情况（默认所有试卷，可按试卷过滤）。"""
    cursor = conn.cursor()
    params = []
    where_clause = "s.status = 'active'"
    if exam_filename:
        where_clause += " AND s.exam_filename = %s"
        params.append(exam_filename)

    cursor.execute(
        f"""
        SELECT
            s.id as session_id,
            s.student_id,
            s.exam_filename,
            st.class_number,
            st.student_number,
            st.name,
            COUNT(a.id) as answered_count,
            SUM(CASE WHEN a.is_correct IS TRUE THEN 1 ELSE 0 END) as correct_count,
            SUM(CASE WHEN a.is_correct IS FALSE THEN 1 ELSE 0 END) as wrong_count,
            COALESCE(SUM(a.score), 0) as total_score
        FROM sessions s
        JOIN students st ON s.student_id = st.id
        LEFT JOIN answers a ON a.session_id = s.id
        WHERE {where_clause}
        GROUP BY s.id, s.student_id, s.exam_filename, st.class_number, st.student_number, st.name
        ORDER BY st.class_number, st.student_number, s.start_time
        """,
        params,
    )
    rows = cursor.fetchall()
    return [dict(row) for row in rows]


@app.get("/api/exams/sidebar_stats")
def get_exam_sidebar_stats(
    exam_filename: Optional[str] = None,
    include_history: bool = True,
    conn: sqlite3.Connection = Depends(get_db),
):
    """返回学生端题号导航侧边栏所需的试卷统计。"""
    current_exam = Path(exam_filename).name if exam_filename else get_current_exam_name(conn) or ""
    if not current_exam:
        return {
            "exam_filename": "",
            "active_sessions": 0,
            "active_students": [],
            "top_students": [],
            "history_included": include_history,
            "history_cached": False,
            "online_refresh_ms": 10000,
            "history_refresh_ms": SIDEBAR_HISTORY_TTL_SECONDS * 1000,
            "health": get_system_health(),
        }

    active_sessions, active_students = load_exam_sidebar_active_students(conn, current_exam)
    health = get_system_health()

    response = {
        "exam_filename": current_exam,
        "active_sessions": active_sessions,
        "active_students": active_students,
        "history_included": include_history,
        "online_refresh_ms": get_sidebar_online_refresh_ms(health),
        "history_refresh_ms": SIDEBAR_HISTORY_TTL_SECONDS * 1000,
        "online_updated_at": datetime.now().isoformat(),
        "health": health,
    }
    if include_history:
        top_students, cached, updated_at = load_exam_sidebar_history(conn, current_exam)
        response["top_students"] = top_students
        response["history_cached"] = cached
        response["history_updated_at"] = updated_at
    return response


@app.get("/api/analysis/export")
def export_exam_answers(
    exam_filename: Optional[str] = None,
    all_exams: bool = False,
    conn: sqlite3.Connection = Depends(get_db),
):
    """导出答题情况（Excel，每次考试一行，含每题得分）。"""
    cursor = conn.cursor()

    def question_sort_key(q):
        qid = q["id"]
        parts = qid.split("-", 1)
        prefix = parts[0]
        num = 0
        if len(parts) == 2:
            try:
                num = int(parts[1])
            except Exception:
                num = 0
        return (prefix, num, qid)

    def write_exam_sheet(wb: Workbook, exam_name: str, title: str | None = None) -> None:
        cursor.execute(
            "SELECT id, type FROM exam_questions WHERE exam_filename = %s",
            (exam_name,),
        )
        question_rows = [dict(r) for r in cursor.fetchall()]
        question_rows.sort(key=question_sort_key)
        question_ids = [q["id"] for q in question_rows]

        cursor.execute(
            """
            SELECT id, student_id, start_time
            FROM sessions
            WHERE exam_filename = %s
            ORDER BY student_id, start_time
            """,
            (exam_name,),
        )
        session_rows = cursor.fetchall()
        attempt_no_by_session = {}
        last_student = None
        count = 0
        for row in session_rows:
            if row["student_id"] != last_student:
                last_student = row["student_id"]
                count = 1
            else:
                count += 1
            attempt_no_by_session[row["id"]] = count

        cursor.execute(
            """
            SELECT 
                s.id as session_id,
                s.student_id,
                s.start_time,
                s.end_time,
                s.total_score,
                s.status,
                s.exam_filename,
                st.class_number,
                st.student_number,
                st.name
            FROM sessions s
            JOIN students st ON s.student_id = st.id
            WHERE s.exam_filename = %s
            ORDER BY st.class_number, st.student_number, s.start_time
            """,
            (exam_name,),
        )
        sessions = [dict(r) for r in cursor.fetchall()]

        cursor.execute(
            """
            SELECT session_id, question_id, score
            FROM answers
            WHERE exam_filename = %s
            """,
            (exam_name,),
        )
        answer_rows = cursor.fetchall()
        scores_by_session = {}
        for row in answer_rows:
            session_id = row["session_id"]
            scores_by_session.setdefault(session_id, {})[row["question_id"]] = (
                row["score"] or 0
            )

        ws = wb.create_sheet(title=title or exam_name[:31])
        headers = [
            "班级",
            "学号",
            "姓名",
            "总分",
            "试卷",
            "考试时间",
            "第几次考试",
        ]
        for idx in range(1, len(question_ids) + 1):
            headers.append(str(idx))
        ws.append(headers)

        for session in sessions:
            session_id = session["session_id"]
            attempt_no = attempt_no_by_session.get(session_id, 1)
            start_time = session.get("start_time")
            end_time = session.get("end_time")
            exam_time = f"{start_time} - {end_time}" if end_time else f"{start_time}"
            scores_map = scores_by_session.get(session_id, {})
            row_values = [
                session["class_number"],
                session["student_number"],
                session["name"],
                session.get("total_score") or 0,
                session.get("exam_filename") or exam_name,
                exam_time,
                attempt_no,
            ]
            for qid in question_ids:
                row_values.append(scores_map.get(qid, 0))
            ws.append(row_values)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16
        ws.column_dimensions["C"].width = 16

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    if all_exams:
        cursor.execute(
            "SELECT DISTINCT exam_filename FROM sessions WHERE exam_filename <> '' ORDER BY exam_filename"
        )
        exam_names = [r[0] for r in cursor.fetchall()]
        if not exam_names:
            raise HTTPException(status_code=400, detail="暂无可导出试卷")
        for name in exam_names:
            write_exam_sheet(wb, name)
        filename = "全部试卷_答题情况.xlsx"
    else:
        current_exam = exam_filename or get_current_exam_name(conn) or ""
        if not current_exam:
            raise HTTPException(status_code=400, detail="当前无可用试卷")
        write_exam_sheet(wb, current_exam, title="答题情况")
        filename = f"{current_exam}_答题情况.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    quoted_name = quote(filename)
    headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{quoted_name}"}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


# 静态文件服务
@app.get("/")
def root():
    return RedirectResponse(url="/xx", status_code=307)


@app.get("/xx")
@app.get("/xx/")
def student_entry():
    return FileResponse(APP_DIR / "index.html", headers={"Cache-Control": "no-store"})


@app.get("/xx/js")
@app.get("/xx/js/")
def teacher_entry():
    return FileResponse(APP_DIR / "teacher.html", headers={"Cache-Control": "no-store"})


@app.get("/xx/student_detail.html")
def teacher_student_detail():
    return FileResponse(
        APP_DIR / "student_detail.html", headers={"Cache-Control": "no-store"}
    )


@app.get("/styles.css")
def serve_styles():
    return FileResponse(BASE_DIR / "styles.css", headers={"Cache-Control": "no-store"})


@app.get("/xx/styles.css")
def serve_styles_xx():
    return FileResponse(BASE_DIR / "styles.css", headers={"Cache-Control": "no-store"})


@app.get("/generated-images/{filename}")
@app.get("/xx/generated-images/{filename}")
def serve_generated_image(filename: str):
    file_path = (GENERATED_IMAGES_DIR / Path(filename).name).resolve()
    if GENERATED_IMAGES_DIR.resolve() in file_path.parents and file_path.exists():
        return FileResponse(file_path, headers={"Cache-Control": "public, max-age=31536000"})
    raise HTTPException(status_code=404, detail="Not Found")


@app.get("/app/{path:path}")
def serve_app(path: str):
    file_path = (APP_DIR / path).resolve()
    if APP_DIR.resolve() in file_path.parents and file_path.exists():
        headers = {"Cache-Control": "no-store"} if file_path.suffix == ".html" else None
        return FileResponse(file_path, headers=headers)
    return {"detail": "Not Found"}


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
