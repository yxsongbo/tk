#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
练习系统后端API
支持100人同时在线练习
技术栈: FastAPI + SQLite
"""

import json
import ast
import re
import sqlite3
import os
from datetime import datetime, timedelta
from typing import Any, List, Optional
from contextlib import asynccontextmanager
from pathlib import Path
import subprocess

from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, RedirectResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import io
from urllib.parse import quote
from pydantic import BaseModel
import uvicorn

# 项目路径
BASE_DIR = Path(__file__).resolve().parent.parent
APP_DIR = BASE_DIR / "app"
EXAMS_DIR = BASE_DIR / "exams"

# 数据库路径
DATABASE = str(BASE_DIR / "practice.db")
DB_BACKEND = "postgres" if os.getenv("DB_BACKEND") == "postgres" else "sqlite"
DATABASE_URL = os.getenv("DATABASE_URL", "")
if DATABASE_URL.startswith("postgres"):
    DB_BACKEND = "postgres"

try:
    import psycopg
    from psycopg.rows import dict_row
    from psycopg_pool import ConnectionPool
except Exception:  # pragma: no cover - optional dependency for Postgres
    psycopg = None
    dict_row = None
    ConnectionPool = None

PG_POOL: Optional["ConnectionPool"] = None


def _is_postgres() -> bool:
    return DB_BACKEND == "postgres"


def _pg_conninfo() -> str:
    if DATABASE_URL:
        return DATABASE_URL
    host = os.getenv("PGHOST", "127.0.0.1")
    port = os.getenv("PGPORT", "5432")
    dbname = os.getenv("PGDATABASE", "tk_practice")
    user = os.getenv("PGUSER", "songbo")
    password = os.getenv("PGPASSWORD", "")
    return (
        f"host={host} port={port} dbname={dbname} user={user} password={password}"
    )


def _fetchone_value(row, default=None):
    if row is None:
        return default
    if isinstance(row, dict):
        return next(iter(row.values()))
    return row[0]


class _SQLiteCompatCursor:
    def __init__(self, cursor: sqlite3.Cursor):
        self._cursor = cursor

    def execute(self, sql: str, params=()):
        return self._cursor.execute(sql.replace("%s", "?"), params)

    def executemany(self, sql: str, seq_of_params):
        return self._cursor.executemany(sql.replace("%s", "?"), seq_of_params)

    def __getattr__(self, name: str):
        return getattr(self._cursor, name)


class _SQLiteCompatConn:
    def __init__(self, conn: sqlite3.Connection):
        self._conn = conn

    def cursor(self):
        return _SQLiteCompatCursor(self._conn.cursor())

    def __getattr__(self, name: str):
        return getattr(self._conn, name)


def normalize_fill_answer(answer: str) -> str:
    """统一填空答案格式：去首尾空白、去全角空格、忽略中间空格。"""
    return str(answer).strip().replace("\u3000", "").replace(" ", "")


def levenshtein_distance(a: str, b: str) -> int:
    """计算两个字符串之间的编辑距离。"""
    if len(a) == 0:
        return len(b)
    if len(b) == 0:
        return len(a)

    matrix = [[0] * (len(b) + 1) for _ in range(len(a) + 1)]

    for i in range(len(a) + 1):
        matrix[i][0] = i
    for j in range(len(b) + 1):
        matrix[0][j] = j

    for i in range(1, len(a) + 1):
        for j in range(1, len(b) + 1):
            if a[i - 1] == b[j - 1]:
                matrix[i][j] = matrix[i - 1][j - 1]
            else:
                matrix[i][j] = min(
                    matrix[i - 1][j - 1] + 1,  # substitution
                    matrix[i][j - 1] + 1,  # insertion
                    matrix[i - 1][j] + 1,  # deletion
                )

    return matrix[len(a)][len(b)]


def strings_similarity(a: str, b: str) -> float:
    """计算两个字符串的相似度（0.0-1.0）。"""
    if a == b:
        return 1.0
    if len(a) == 0 or len(b) == 0:
        return 0.0

    distance = levenshtein_distance(a, b)
    max_length = max(len(a), len(b))
    return 1.0 - (distance / max_length)


def build_fill_answer_variants(answer: str) -> set[str]:
    """从标准答案中提取可接受答案（支持“或”前后写法）。"""
    raw = str(answer).strip()
    if not raw:
        return set()

    variants = {normalize_fill_answer(raw)}

    normalized_brackets = raw.replace("（", "(").replace("）", ")")

    # 兼容：A（或B）/ A(或B) / A或B / A，或B / A,或B
    if "或" in normalized_brackets:
        parts = re.split(r"\s*[,，]?\s*或\s*", normalized_brackets)
        for part in parts:
            cleaned = part.strip().strip("()[]{}<>\"'“”‘’、，,;；")
            candidate = normalize_fill_answer(cleaned)
            if candidate:
                variants.add(candidate)

            # 兼容：字符串(str) 这种混合写法
            m = re.match(r"^(.*?)\((.*?)\)$", cleaned)
            if m:
                left = normalize_fill_answer(m.group(1))
                right = normalize_fill_answer(m.group(2))
                if left:
                    variants.add(left)
                if right:
                    variants.add(right)

    return {v for v in variants if v}


def is_fill_answer_match(student_answer: str, correct_answer: str) -> float:
    """填空匹配：含“或”时，允许匹配“或”前后内容（含近似包含）。返回0.0-1.0的相似度，>=0.9返回1.0（满分）。"""
    student = normalize_fill_answer(student_answer)
    if not student:
        return 0.0

    raw = str(correct_answer or "").strip()
    variants = build_fill_answer_variants(raw)
    if not variants:
        return 0.0

    max_similarity = 0.0

    # Check exact match first
    if student in variants:
        return 1.0

    # Calculate similarity with all variants
    for variant in variants:
        similarity = strings_similarity(student, variant)
        if similarity > max_similarity:
            max_similarity = similarity

    # For answers containing "或", also check substring relationship
    if "或" in raw.replace("（", "(").replace("）", ")"):
        for variant in variants:
            if (
                len(student) >= 2
                and len(variant) >= 2
                and (student in variant or variant in student)
            ):
                # Calculate similarity for substring relationship
                shorter = student if len(student) < len(variant) else variant
                longer = student if len(student) >= len(variant) else variant
                similarity = strings_similarity(shorter, longer)
                if similarity > max_similarity:
                    max_similarity = similarity

    # If similarity >= 90%, return full score
    if max_similarity >= 0.9:
        return 1.0

    return max_similarity


def _coerce_str_list(value: Any) -> list[str]:
    """把值统一转换为字符串数组（去掉空值）。"""
    if value is None:
        return []
    if isinstance(value, list):
        values = []
        for item in value:
            if item is None:
                continue
            text = str(item).strip()
            if text:
                values.append(text)
        return values
    text = str(value).strip()
    return [text] if text else []


FORMULA_FRAGMENT_RE = re.compile(r"(=|\$|\(|\)|:|[A-Za-z]+\$?\d+)")


def _dedupe_preserve_order(items: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        if item not in seen:
            seen.add(item)
            result.append(item)
    return result


def _should_join_answer_fragments(answers: list[str]) -> bool:
    if len(answers) < 2:
        return False
    if any("或" in answer for answer in answers):
        return False
    if any("," in answer or "，" in answer for answer in answers):
        return False
    combined = ",".join(answer.strip() for answer in answers if answer)
    if not combined:
        return False
    if combined.startswith("="):
        return True
    return bool(FORMULA_FRAGMENT_RE.search(combined))


def _expand_fill_blank_answers(answers: list[str]) -> list[str]:
    cleaned = [str(item).strip() for item in answers if str(item).strip()]
    if not cleaned:
        return []
    cleaned = _dedupe_preserve_order(cleaned)
    if _should_join_answer_fragments(cleaned):
        for joiner in (",", "，"):
            joined = joiner.join(cleaned)
            if joined and joined not in cleaned:
                cleaned.append(joined)
    return cleaned


def _coerce_image_list(value: Any) -> list[str]:
    """把单图/多图字段统一为字符串数组。"""
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return []
        if raw.startswith("["):
            try:
                parsed = json.loads(raw)
                if isinstance(parsed, list):
                    return [str(v).strip() for v in parsed if str(v).strip()]
            except Exception:
                pass
        return [raw]
    text = str(value).strip()
    return [text] if text else []


def _normalize_fill_blank_item(item: Any) -> dict:
    """标准化填空每一空的结构。"""
    if isinstance(item, dict):
        answers = _expand_fill_blank_answers(_coerce_str_list(item.get("answers")))
        images = _coerce_image_list(item.get("images"))
        if not images:
            images = _coerce_image_list(item.get("image"))
        label = item.get("label")
        label_text = str(label).strip() if label is not None else None
        if label_text == "":
            label_text = None
        return {"answers": answers, "images": images, "label": label_text}

    if isinstance(item, list):
        return {
            "answers": _expand_fill_blank_answers(_coerce_str_list(item)),
            "images": [],
            "label": None,
        }

    return {
        "answers": _expand_fill_blank_answers(_coerce_str_list(item)),
        "images": [],
        "label": None,
    }


def parse_fill_blanks(raw: Any) -> list[dict]:
    """
    解析填空答案，兼容多种历史格式：
    1) JSON数组/对象字符串
    2) Python 字面量字符串（旧数据）
    3) 逗号分隔字符串
    """
    if raw is None:
        return []

    parsed: Any = raw
    if isinstance(raw, str):
        text = raw.strip()
        if not text:
            return []

        parsed = text
        parse_candidates: list[str] = []
        if text.startswith("[") or text.startswith("{"):
            parse_candidates.append(text)
        parse_candidates.append(text)
        parse_candidates.append(f"[{text}]")

        for candidate in parse_candidates:
            try:
                parsed = json.loads(candidate)
                break
            except Exception:
                try:
                    parsed = ast.literal_eval(candidate)
                    break
                except Exception:
                    parsed = text

        if isinstance(parsed, str):
            # Treat the entire string as a single acceptable answer for one blank
            # Don't split on commas, as they may be part of the answer (e.g., Excel formulas)
            return [{"answers": [text], "images": [], "label": None}]

    if isinstance(parsed, dict):
        if isinstance(parsed.get("blanks"), list):
            parsed = parsed.get("blanks")
        elif isinstance(parsed.get("answers"), list):
            parsed = parsed.get("answers")
        else:
            return []

    if isinstance(parsed, list):
        blanks = [_normalize_fill_blank_item(item) for item in parsed]
        return [
            blank
            for blank in blanks
            if blank["answers"] or blank["images"] or blank["label"]
        ]

    return [_normalize_fill_blank_item(parsed)]


def ensure_column(cursor: Any, table: str, column: str, ddl: str) -> None:
    """为已有表补充字段（若不存在）。"""
    if _is_postgres():
        cursor.execute(
            "SELECT column_name FROM information_schema.columns WHERE table_name = %s",
            (table,),
        )
        rows = cursor.fetchall()
        existing_columns = {
            (row["column_name"] if isinstance(row, dict) else row[0]) for row in rows
        }
    else:
        cursor.execute(f"PRAGMA table_info({table})")
        existing_columns = {row[1] for row in cursor.fetchall()}
    if column not in existing_columns:
        cursor.execute(f"ALTER TABLE {table} ADD COLUMN {ddl}")


def ensure_exam_questions_table(cursor: sqlite3.Cursor) -> None:
    """确保 exam_questions 表存在，用于缓存多份试卷的题目。"""
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS exam_questions (
            exam_filename TEXT NOT NULL,
            id TEXT NOT NULL,
            type TEXT NOT NULL CHECK(type IN ('choice', 'fill')),
            question TEXT NOT NULL,
            score REAL NOT NULL,
            explanation TEXT,
            image TEXT,
            correct_answer TEXT,
            options TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (exam_filename, id)
        )
        """
    )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_exam_questions_exam ON exam_questions(exam_filename)"
    )


def read_meminfo() -> dict[str, int]:
    """读取 /proc/meminfo，返回 kB 数值字典；失败则返回空。"""
    info: dict[str, int] = {}
    try:
        with open("/proc/meminfo", "r") as f:
            for line in f:
                parts = line.split(":")
                if len(parts) >= 2:
                    key = parts[0].strip()
                    value = parts[1].strip().split()[0]
                    info[key] = int(value)
    except Exception:
        pass
    return info


def read_swaps() -> tuple[int, int]:
    """读取 /proc/swaps，返回 (total_kb, used_kb)。"""
    total = used = 0
    try:
        with open("/proc/swaps", "r") as f:
            lines = f.readlines()[1:]  # skip header
            for line in lines:
                parts = line.split()
                if len(parts) >= 5:
                    total += int(parts[2])
                    used += int(parts[3])
    except Exception:
        pass
    return total, used


def get_system_health() -> dict:
    """轻量级系统负载指标，避免额外依赖。"""
    meminfo = read_meminfo()
    mem_avail_kb = meminfo.get("MemAvailable", 0)
    mem_total_kb = meminfo.get("MemTotal", 0)
    swap_total_kb, swap_used_kb = read_swaps()
    load1, load5, load15 = os.getloadavg() if hasattr(os, "getloadavg") else (0, 0, 0)

    # 简单告警判定：可用内存 < 80MB 或 swap 使用率 > 80% 或 1分钟负载 > CPU核数*1.5
    cpu_count = os.cpu_count() or 1
    mem_alert = mem_avail_kb < 80 * 1024
    swap_alert = swap_total_kb > 0 and (swap_used_kb / swap_total_kb) > 0.8
    load_alert = load1 > cpu_count * 1.5
    degraded = mem_alert or swap_alert or load_alert

    return {
        "mem_available_mb": round(mem_avail_kb / 1024, 1),
        "mem_total_mb": round(mem_total_kb / 1024, 1),
        "swap_total_mb": round(swap_total_kb / 1024, 1),
        "swap_used_mb": round(swap_used_kb / 1024, 1),
        "load1": round(load1, 2),
        "load5": round(load5, 2),
        "load15": round(load15, 2),
        "cpu_count": cpu_count,
        "alert": degraded,
        "alert_reasons": {
            "low_memory": mem_alert,
            "high_swap": swap_alert,
            "high_load": load_alert,
        },
    }


def ensure_exam_questions_table(cursor: sqlite3.Cursor) -> None:
    """确保 exam_questions 表存在，用于缓存多份试卷的题目。"""
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS exam_questions (
            exam_filename TEXT NOT NULL,
            id TEXT NOT NULL,
            type TEXT NOT NULL CHECK(type IN ('choice', 'fill')),
            question TEXT NOT NULL,
            score REAL NOT NULL,
            explanation TEXT,
            image TEXT,
            correct_answer TEXT,
            options TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (exam_filename, id)
        )
        """
    )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_exam_questions_exam ON exam_questions(exam_filename)"
    )


def extract_correct_answer(question_data: dict) -> str | int | None:
    """兼容不同试卷格式的答案字段。"""
    question_type = str(question_data.get("type", "")).strip()

    if question_type == "fill":
        fill_blanks = parse_fill_blanks(question_data.get("answers"))
        if not fill_blanks:
            fill_blanks = parse_fill_blanks(question_data.get("correctAnswer"))
        if fill_blanks:
            return json.dumps(fill_blanks, ensure_ascii=False)
        return None

    correct_answer = question_data.get("correctAnswer")
    if correct_answer is None and "answers" in question_data:
        answers = question_data.get("answers", [])
        if answers:
            all_answers = []
            for answer_group in answers:
                if isinstance(answer_group, list):
                    all_answers.extend(
                        [str(a) for a in answer_group if a is not None and a != ""]
                    )
                elif answer_group is not None and answer_group != "":
                    all_answers.append(str(answer_group))
            correct_answer = ",".join(all_answers) if all_answers else None
    return correct_answer


def normalize_image_field(image_value):
    """兼容单图/多图：数据库中 image 可能是字符串或 JSON 数组字符串。"""
    if image_value is None:
        return None
    if isinstance(image_value, list):
        return [str(v) for v in image_value if v]
    if not isinstance(image_value, str):
        return str(image_value)

    raw = image_value.strip()
    if not raw:
        return None

    if raw.startswith("["):
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                return [str(v) for v in parsed if v]
        except Exception:
            pass
    return raw


def list_exam_files() -> list[Path]:
    """列出可用试卷文件（根目录 + exams 子目录）。"""
    EXAMS_DIR.mkdir(parents=True, exist_ok=True)
    files_by_name: dict[str, Path] = {}

    for exam_path in sorted(BASE_DIR.glob("*.json")):
        files_by_name[exam_path.name] = exam_path
    for exam_path in sorted(EXAMS_DIR.glob("*.json")):
        files_by_name[exam_path.name] = exam_path

    return [files_by_name[name] for name in sorted(files_by_name.keys())]


def resolve_exam_path(filename: str) -> Path | None:
    """按文件名解析试卷路径。"""
    safe_name = Path(filename).name
    for exam_path in list_exam_files():
        if exam_path.name == safe_name:
            return exam_path
    return None


def load_exam_json(exam_path: Path) -> dict:
    """读取并校验试卷 JSON。"""
    try:
        with open(exam_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"读取试卷失败: {e}") from e

    questions = data.get("questions")
    if not isinstance(questions, list) or not questions:
        raise HTTPException(
            status_code=400, detail="试卷格式错误：questions 必须为非空数组"
        )
    return data


def set_current_exam(conn: sqlite3.Connection, filename: str) -> None:
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO settings (key, value, updated_at)
        VALUES ('current_exam', %s, CURRENT_TIMESTAMP)
        ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = CURRENT_TIMESTAMP
        """,
        (filename,),
    )


def get_current_exam_name(conn: sqlite3.Connection) -> str | None:
    cursor = conn.cursor()
    cursor.execute("SELECT value FROM settings WHERE key = 'current_exam'")
    row = cursor.fetchone()
    if row and row["value"]:
        return row["value"]
    return None


def get_exam_mode(conn: sqlite3.Connection) -> str:
    """获取当前模式：practice 或 exam。默认 practice。"""
    cursor = conn.cursor()
    cursor.execute("SELECT value FROM settings WHERE key = 'exam_mode'")
    row = cursor.fetchone()
    if row and row["value"] in ("practice", "exam"):
        return row["value"]
    return "practice"


def set_exam_mode(conn: sqlite3.Connection, mode: str) -> None:
    if mode not in ("practice", "exam"):
        raise HTTPException(status_code=400, detail="mode 必须为 practice 或 exam")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO settings (key, value, updated_at)
        VALUES ('exam_mode', %s, CURRENT_TIMESTAMP)
        ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = CURRENT_TIMESTAMP
        """,
        (mode,),
    )


def get_exam_distribution(conn: sqlite3.Connection) -> tuple[str, Optional[str]]:
    """
    返回考试分发策略:
    - ('random', None): 随机分发现有试卷
    - ('fixed', filename): 固定分发指定试卷
    """
    cursor = conn.cursor()
    cursor.execute("SELECT value FROM settings WHERE key = 'exam_distribution'")
    dist = cursor.fetchone()
    cursor.execute("SELECT value FROM settings WHERE key = 'exam_fixed'")
    fixed = cursor.fetchone()
    dist_value = dist["value"] if dist and dist["value"] in ("random", "fixed") else "random"
    fixed_value = fixed["value"] if fixed and fixed["value"] else None
    return dist_value, fixed_value


def set_exam_distribution(conn: sqlite3.Connection, mode: str, filename: Optional[str]) -> None:
    if mode not in ("random", "fixed"):
        raise HTTPException(status_code=400, detail="distribution 必须为 random 或 fixed")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO settings (key, value, updated_at)
        VALUES ('exam_distribution', %s, CURRENT_TIMESTAMP)
        ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = CURRENT_TIMESTAMP
        """,
        (mode,),
    )
    cursor.execute(
        """
        INSERT INTO settings (key, value, updated_at)
        VALUES ('exam_fixed', %s, CURRENT_TIMESTAMP)
        ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = CURRENT_TIMESTAMP
        """,
        (filename or ""),
    )


def import_exam_into_questions(conn: sqlite3.Connection, exam_path: Path) -> dict:
    """把选中的试卷导入 questions 表（覆盖当前题库）。"""
    data = load_exam_json(exam_path)
    questions = data["questions"]
    cursor = conn.cursor()

    # 同步到全局 questions（兼容旧前端/教师端）
    cursor.execute("DELETE FROM questions")

    insert_count = 0
    for q in questions:
        if not q.get("id") or not q.get("type") or not q.get("question"):
            raise HTTPException(status_code=400, detail=f"试卷题目字段缺失: {q}")

        options = (
            json.dumps(q.get("options", []), ensure_ascii=False)
            if "options" in q
            else None
        )
        correct_answer = extract_correct_answer(q)
        image_value = q.get("image")
        if not image_value and isinstance(q.get("images"), list):
            image_value = json.dumps(q.get("images", []), ensure_ascii=False)

        cursor.execute(
            """
            INSERT INTO questions (id, type, question, score, explanation, image, correct_answer, options)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                str(q["id"]),
                str(q["type"]),
                str(q["question"]),
                float(q.get("score", 0)),
                q.get("explanation", ""),
                image_value,
                correct_answer,
                options,
            ),
        )
        insert_count += 1

    # 也缓存到 exam_questions（用于多试卷并行）
    import_exam_into_exam_questions(conn, exam_path, data=data)

    set_current_exam(conn, exam_path.name)
    conn.commit()

    return {
        "filename": exam_path.name,
        "title": data.get("title", exam_path.name),
        "question_count": insert_count,
    }


def import_exam_into_exam_questions(
    conn: sqlite3.Connection, exam_path: Path, data: Optional[dict] = None
) -> dict:
    """把试卷题目写入 exam_questions（不影响当前 questions）。"""
    exam_data = data or load_exam_json(exam_path)
    questions = exam_data["questions"]
    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM exam_questions WHERE exam_filename = %s", (exam_path.name,)
    )

    count = 0
    for q in questions:
        if not q.get("id") or not q.get("type") or not q.get("question"):
            raise HTTPException(status_code=400, detail=f"试卷题目字段缺失: {q}")
        options = (
            json.dumps(q.get("options", []), ensure_ascii=False)
            if "options" in q
            else None
        )
        correct_answer = extract_correct_answer(q)
        image_value = q.get("image")
        if not image_value and isinstance(q.get("images"), list):
            image_value = json.dumps(q.get("images", []), ensure_ascii=False)

        cursor.execute(
            """
            INSERT INTO exam_questions (exam_filename, id, type, question, score, explanation, image, correct_answer, options)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                exam_path.name,
                str(q["id"]),
                str(q["type"]),
                str(q["question"]),
                float(q.get("score", 0)),
                q.get("explanation", ""),
                image_value,
                correct_answer,
                options,
            ),
        )
        count += 1

    conn.commit()
    return {
        "filename": exam_path.name,
        "title": exam_data.get("title", exam_path.name),
        "question_count": count,
    }


# 数据模型
class Student(BaseModel):
    id: int
    exam_number: str
    class_number: int
    student_number: int
    name: str
    subject_group: int


class Question(BaseModel):
    id: str
    type: str
    question: str
    score: float
    explanation: Optional[str] = None
    image: Optional[str | List[str]] = None
    correct_answer: Optional[str] = None
    fill_blanks: Optional[List[dict]] = None
    options: Optional[List[dict]] = None


class AnswerRequest(BaseModel):
    student_id: int
    session_id: int
    question_id: str
    answer: str | int
    answer_time: int  # 答题用时（秒）


class SessionStart(BaseModel):
    student_id: int
    exam_filename: Optional[str] = None
    override_code: Optional[str] = None


class SessionResponse(BaseModel):
    id: int
    student_id: int
    start_time: str
    status: str
    exam_filename: Optional[str] = None
    resumed: Optional[bool] = None


# 数据库连接
def get_db():
    if _is_postgres():
        if psycopg is None:
            raise RuntimeError("psycopg 未安装，无法连接 PostgreSQL")
        global PG_POOL
        if PG_POOL is None:
            if ConnectionPool is None:
                raise RuntimeError("psycopg_pool 未安装，无法创建连接池")
            PG_POOL = ConnectionPool(
                _pg_conninfo(),
                min_size=2,
                max_size=30,
                kwargs={"row_factory": dict_row},
            )
        with PG_POOL.connection() as conn:
            yield conn
    else:
        raw_conn = sqlite3.connect(DATABASE, check_same_thread=False, timeout=30)
        raw_conn.row_factory = sqlite3.Row
        raw_conn.execute("PRAGMA busy_timeout = 30000")
        # Keep SQLite settings consistent per-connection under concurrent load
        raw_conn.execute("PRAGMA journal_mode=WAL")
        raw_conn.execute("PRAGMA synchronous=NORMAL")
        raw_conn.execute("PRAGMA temp_store=MEMORY")
        conn = _SQLiteCompatConn(raw_conn)
        try:
            yield conn
        finally:
            raw_conn.close()


# 初始化数据库
@asynccontextmanager
async def lifespan(app: FastAPI):
    # 启动时执行
    init_db()
    yield
    # 关闭时执行
    global PG_POOL
    if PG_POOL is not None:
        PG_POOL.close()


app = FastAPI(title="练习系统API", lifespan=lifespan)

# CORS配置
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.api_route(
    "/xx/api/{path:path}",
    methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS", "HEAD"],
)
def local_xx_api_redirect(path: str, request: Request):
    """
    本地直跑 FastAPI 时，前端会请求 /xx/api/*。
    线上由 Nginx 重写到 /api/*；本地没有 Nginx，因此这里做兼容跳转。
    """
    target = f"/api/{path}"
    if request.url.query:
        target = f"{target}?{request.url.query}"
    return RedirectResponse(url=target, status_code=307)


def init_db_postgres():
    """初始化 PostgreSQL 数据库"""
    if psycopg is None:
        raise RuntimeError("psycopg 未安装，无法初始化 PostgreSQL")
    conn = psycopg.connect(_pg_conninfo(), row_factory=dict_row)
    cursor = conn.cursor()

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS students (
            id SERIAL PRIMARY KEY,
            exam_number TEXT UNIQUE NOT NULL,
            class_number INTEGER NOT NULL,
            student_number INTEGER NOT NULL,
            name TEXT NOT NULL,
            subject_group INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS questions (
            id TEXT PRIMARY KEY,
            type TEXT NOT NULL CHECK(type IN ('choice', 'fill')),
            question TEXT NOT NULL,
            score REAL NOT NULL,
            explanation TEXT,
            image TEXT,
            correct_answer TEXT,
            options TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS sessions (
            id SERIAL PRIMARY KEY,
            student_id INTEGER NOT NULL REFERENCES students(id),
            start_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            end_time TIMESTAMP,
            total_score REAL DEFAULT 0,
            status TEXT DEFAULT 'active' CHECK(status IN ('active', 'completed', 'abandoned')),
            exam_filename TEXT DEFAULT ''
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS answers (
            id SERIAL PRIMARY KEY,
            session_id INTEGER NOT NULL REFERENCES sessions(id),
            student_id INTEGER NOT NULL REFERENCES students(id),
            question_id TEXT NOT NULL,
            answer TEXT,
            is_correct BOOLEAN,
            score REAL DEFAULT 0,
            answer_time INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            exam_filename TEXT DEFAULT '',
            UNIQUE(session_id, question_id)
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )

    ensure_exam_questions_table(cursor)
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_answers_question ON answers(question_id)"
    )
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_answers_student ON answers(student_id)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_answers_session ON answers(session_id)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_answers_created ON answers(created_at)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_students_exam ON students(exam_number)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_sessions_exam ON sessions(exam_filename)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_answers_exam ON answers(exam_filename)")
    cursor.execute(
        "ALTER TABLE answers DROP CONSTRAINT IF EXISTS answers_question_id_fkey"
    )

    cursor.execute(
        "INSERT INTO settings (key, value) VALUES ('exam_mode', 'practice') ON CONFLICT (key) DO NOTHING"
    )
    cursor.execute(
        "INSERT INTO settings (key, value) VALUES ('exam_distribution', 'random') ON CONFLICT (key) DO NOTHING"
    )
    cursor.execute(
        "INSERT INTO settings (key, value) VALUES ('exam_fixed', '') ON CONFLICT (key) DO NOTHING"
    )

    # 启动时保证当前试卷可用
    current_exam = get_current_exam_name(conn)
    available_exams = list_exam_files()
    if current_exam:
        current_exam_path = resolve_exam_path(current_exam)
        if current_exam_path:
            import_exam_into_questions(conn, current_exam_path)
            import_exam_into_exam_questions(conn, current_exam_path)
        elif available_exams:
            import_exam_into_questions(conn, available_exams[0])
            import_exam_into_exam_questions(conn, available_exams[0])
        else:
            set_current_exam(conn, "")
    elif available_exams:
        import_exam_into_questions(conn, available_exams[0])
        import_exam_into_exam_questions(conn, available_exams[0])

    conn.commit()
    conn.close()


def init_db():
    """初始化数据库"""
    if _is_postgres():
        init_db_postgres()
        return
    conn = sqlite3.connect(DATABASE, check_same_thread=False, timeout=30)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("PRAGMA journal_mode=WAL")
    cursor.execute("PRAGMA synchronous=NORMAL")

    # 检查表是否已存在
    cursor.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='students'"
    )
    if not cursor.fetchone():
        # 表不存在，执行建表脚本
        with open(BASE_DIR / "database" / "schema.sql", "r", encoding="utf-8") as f:
            conn.executescript(f.read())
        print("数据库初始化完成")
    else:
        print("数据库已存在，跳过初始化")

    # 确保settings表存在
    cursor.execute(
        "CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)"
    )
    cursor.execute(
        "INSERT OR IGNORE INTO settings (key, value) VALUES ('exam_mode', 'practice')"
    )
    cursor.execute(
        "INSERT OR IGNORE INTO settings (key, value) VALUES ('exam_distribution', 'random')"
    )
    cursor.execute(
        "INSERT OR IGNORE INTO settings (key, value) VALUES ('exam_fixed', '')"
    )

    # 历史库结构兼容：补充试卷字段
    ensure_column(cursor, "sessions", "exam_filename", "exam_filename TEXT")
    ensure_column(cursor, "answers", "exam_filename", "exam_filename TEXT")
    ensure_exam_questions_table(cursor)
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_sessions_exam ON sessions(exam_filename)"
    )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_answers_exam ON answers(exam_filename)"
    )
    cursor.execute("UPDATE sessions SET exam_filename = '' WHERE exam_filename IS NULL")
    cursor.execute("UPDATE answers SET exam_filename = '' WHERE exam_filename IS NULL")

    # 启动时保证当前试卷可用
    current_exam = get_current_exam_name(conn)
    available_exams = list_exam_files()
    if current_exam:
        current_exam_path = resolve_exam_path(current_exam)
        if current_exam_path:
            import_exam_into_questions(conn, current_exam_path)
            import_exam_into_exam_questions(conn, current_exam_path)
        elif available_exams:
            import_exam_into_questions(conn, available_exams[0])
            import_exam_into_exam_questions(conn, available_exams[0])
        else:
            set_current_exam(conn, "")
    elif available_exams:
        import_exam_into_questions(conn, available_exams[0])
        import_exam_into_exam_questions(conn, available_exams[0])

    conn.commit()
    conn.close()


# ========== 学生相关接口 ==========


@app.get("/api/students", response_model=List[Student])
def get_students(
    class_number: Optional[int] = None, conn: sqlite3.Connection = Depends(get_db)
):
    """获取学生列表"""
    cursor = conn.cursor()
    if class_number:
        cursor.execute("SELECT * FROM students WHERE class_number = %s", (class_number,))
    else:
        cursor.execute("SELECT * FROM students ORDER BY class_number, student_number")
    rows = cursor.fetchall()
    return [Student(**dict(row)) for row in rows]


@app.get("/api/students/{student_id}", response_model=Student)
def get_student(student_id: int, conn: sqlite3.Connection = Depends(get_db)):
    """获取单个学生信息"""
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM students WHERE id = %s", (student_id,))
    row = cursor.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="学生不存在")
    return Student(**dict(row))


# ========== 题目相关接口 ==========


@app.get("/api/questions", response_model=List[Question])
def get_questions(
    type: Optional[str] = None,
    exam_filename: Optional[str] = None,
    conn: sqlite3.Connection = Depends(get_db),
):
    """获取题目列表"""
    try:
        exam_mode = get_exam_mode(conn)
        cursor = conn.cursor()
        rows = []
        if exam_filename:
            cursor.execute(
                "SELECT * FROM exam_questions WHERE exam_filename = %s",
                (Path(exam_filename).name,),
            )
            rows = cursor.fetchall()
        else:
            if type:
                cursor.execute("SELECT * FROM questions WHERE type = %s", (type,))
            else:
                cursor.execute("SELECT * FROM questions ORDER BY id")
            rows = cursor.fetchall()

        questions = []
        for row in rows:
            q = dict(row)
            if q["options"]:
                q["options"] = json.loads(q["options"])
            q["image"] = normalize_image_field(q.get("image"))
            if exam_mode == "practice":
                if q.get("correct_answer") is not None:
                    q["correct_answer"] = str(q["correct_answer"])
                if q.get("type") == "fill":
                    q["fill_blanks"] = parse_fill_blanks(q.get("correct_answer"))
            else:
                blanks = (
                    parse_fill_blanks(q.get("correct_answer"))
                    if q.get("type") == "fill"
                    else []
                )
                for blank in blanks:
                    blank["answers"] = []
                q["correct_answer"] = None
                q["fill_blanks"] = blanks
                q["explanation"] = None
            questions.append(Question(**q))
        return questions
    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/questions/{question_id}", response_model=Question)
def get_question(
    question_id: str,
    exam_filename: Optional[str] = None,
    conn: sqlite3.Connection = Depends(get_db),
):
    """获取单道题目"""
    cursor = conn.cursor()
    exam_mode = get_exam_mode(conn)
    if exam_filename:
        cursor.execute(
            "SELECT * FROM exam_questions WHERE exam_filename = %s AND id = %s",
            (Path(exam_filename).name, question_id),
        )
        row = cursor.fetchone()
    else:
        cursor.execute("SELECT * FROM questions WHERE id = %s", (question_id,))
        row = cursor.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="题目不存在")

    q = dict(row)
    if q["options"]:
        q["options"] = json.loads(q["options"])
    q["image"] = normalize_image_field(q.get("image"))
    if exam_mode == "practice":
        if q.get("correct_answer") is not None:
            q["correct_answer"] = str(q["correct_answer"])
        if q.get("type") == "fill":
            q["fill_blanks"] = parse_fill_blanks(q.get("correct_answer"))
    else:
        blanks = (
            parse_fill_blanks(q.get("correct_answer"))
            if q.get("type") == "fill"
            else []
        )
        for blank in blanks:
            blank["answers"] = []
        q["correct_answer"] = None
        q["fill_blanks"] = blanks
        q["explanation"] = None
    return Question(**q)


# ========== 练习会话接口 ==========


@app.post("/api/sessions/start", response_model=SessionResponse)
def start_session(data: SessionStart, conn: sqlite3.Connection = Depends(get_db)):
    """开始练习会话"""
    cursor = conn.cursor()
    exam_mode = get_exam_mode(conn)

    # 如果已有进行中的会话，优先恢复（除非提供口令强制重开）
    cursor.execute(
        """
        SELECT id, student_id, start_time, status, exam_filename
        FROM sessions
        WHERE student_id = %s AND status = 'active'
        ORDER BY start_time DESC
        LIMIT 1
        """,
        (data.student_id,),
    )
    existing = cursor.fetchone()
    if existing and data.override_code != "2055350":
        existing_exam = (existing["exam_filename"] or "").strip()
        if not existing_exam:
            existing_exam = get_current_exam_name(conn) or ""
            cursor.execute(
                "UPDATE sessions SET exam_filename = %s WHERE id = %s",
                (existing_exam, existing["id"]),
            )
            conn.commit()
        if existing_exam:
            cursor.execute(
                "SELECT COUNT(*) FROM exam_questions WHERE exam_filename = %s",
                (existing_exam,),
            )
        if (_fetchone_value(cursor.fetchone(), 0) or 0) == 0:
                path = resolve_exam_path(existing_exam)
                if path:
                    import_exam_into_exam_questions(conn, path)
        return SessionResponse(
            id=existing["id"],
            student_id=existing["student_id"],
            start_time=str(existing["start_time"]),
            status=existing["status"],
            exam_filename=existing_exam,
            resumed=True,
        )

    # 确定本次会话使用的试卷
    available_exams = list_exam_files()
    if not available_exams:
        raise HTTPException(status_code=400, detail="当前无可用试卷，请先在教师端导入")

    if exam_mode == "practice":
        # 学生自选试卷（未传则使用当前/首个）
        chosen = data.exam_filename or get_current_exam_name(conn)
        if not chosen and available_exams:
            chosen = available_exams[0].name
        exam_filename = Path(chosen).name if chosen else None
    else:
        # 考试模式：按策略分发
        dist, fixed = get_exam_distribution(conn)
        if dist == "fixed" and fixed:
            exam_filename = fixed
        else:
            # 随机分发现有试卷
            import random

            exam_filename = random.choice(available_exams).name

    if not exam_filename or not resolve_exam_path(exam_filename):
        raise HTTPException(status_code=404, detail="试卷文件不存在")

    # 所有模式：1小时内仅允许一次会话，除非提供口令
    one_hour_ago = datetime.now() - timedelta(hours=1)
    cursor.execute(
        """
        SELECT COUNT(*) FROM sessions
        WHERE student_id = %s AND start_time >= %s AND status IN ('active', 'completed')
        """,
        (data.student_id, one_hour_ago),
    )
    recent_count = _fetchone_value(cursor.fetchone(), 0) or 0
    if recent_count > 0 and data.override_code != "2055350":
        raise HTTPException(
            status_code=429,
            detail="每小时仅允许登录一次。如需重登请联系教师获取口令。",
        )

    # 确保 exam_questions 有缓存，没有则导入
    cursor.execute(
        "SELECT COUNT(*) FROM exam_questions WHERE exam_filename = %s",
        (exam_filename,),
    )
    if (_fetchone_value(cursor.fetchone(), 0) or 0) == 0:
        path = resolve_exam_path(exam_filename)
        if not path:
            raise HTTPException(status_code=404, detail="试卷文件不存在")
        import_exam_into_exam_questions(conn, path)

    # 检查是否有进行中的会话（口令强制重开时才会走到这里）
    cursor.execute(
        "SELECT id FROM sessions WHERE student_id = %s AND status = 'active'",
        (data.student_id,),
    )
    existing = cursor.fetchone()
    if existing:
        cursor.execute(
            "UPDATE sessions SET status = 'abandoned', end_time = %s WHERE id = %s",
            (datetime.now(), existing["id"]),
        )

    # 创建新会话
    if _is_postgres():
        cursor.execute(
            "INSERT INTO sessions (student_id, start_time, status, exam_filename) VALUES (%s, %s, 'active', %s) RETURNING id",
            (data.student_id, datetime.now(), exam_filename),
        )
        session_id = _fetchone_value(cursor.fetchone())
    else:
        cursor.execute(
            "INSERT INTO sessions (student_id, start_time, status, exam_filename) VALUES (%s, %s, 'active', %s)",
            (data.student_id, datetime.now(), exam_filename),
        )
        session_id = cursor.lastrowid
    conn.commit()
    return SessionResponse(
        id=session_id,
        student_id=data.student_id,
        start_time=datetime.now().isoformat(),
        status="active",
        exam_filename=exam_filename,
        resumed=False,
    )


@app.get("/api/sessions/{session_id}/state")
def get_session_state(
    session_id: int, conn: sqlite3.Connection = Depends(get_db)
):
    """获取会话已答题目与得分，用于恢复进度。"""
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, student_id, status, exam_filename FROM sessions WHERE id = %s",
        (session_id,),
    )
    session = cursor.fetchone()
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")
    if session["status"] != "active":
        raise HTTPException(status_code=400, detail="会话已结束，不能恢复")

    cursor.execute(
        """
        SELECT question_id, answer, is_correct, score
        FROM answers
        WHERE session_id = %s
        ORDER BY created_at ASC
        """,
        (session_id,),
    )
    answers = [dict(row) for row in cursor.fetchall()]
    total_score = sum(float(a.get("score") or 0) for a in answers)

    return {
        "session_id": session_id,
        "exam_filename": (session["exam_filename"] or "").strip(),
        "answers": answers,
        "total_score": total_score,
    }


@app.post("/api/sessions/{session_id}/end")
def end_session(session_id: int, conn: sqlite3.Connection = Depends(get_db)):
    """结束练习会话"""
    cursor = conn.cursor()

    # 计算总分
    cursor.execute("SELECT SUM(score) FROM answers WHERE session_id = %s", (session_id,))
    total_score = _fetchone_value(cursor.fetchone(), 0) or 0

    cursor.execute(
        "UPDATE sessions SET end_time = %s, total_score = %s, status = 'completed' WHERE id = %s",
        (datetime.now(), total_score, session_id),
    )
    conn.commit()

    return {"message": "练习已结束", "total_score": total_score}


# ========== 答题接口 ==========


@app.post("/api/answers/submit")
def submit_answer(data: AnswerRequest, conn: sqlite3.Connection = Depends(get_db)):
    """提交答案"""
    cursor = conn.cursor()
    cursor.execute(
        "SELECT status, exam_filename FROM sessions WHERE id = %s AND student_id = %s",
        (data.session_id, data.student_id),
    )
    session = cursor.fetchone()
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")
    if session["status"] != "active":
        raise HTTPException(status_code=400, detail="会话已结束，不能继续答题")

    session_exam = session["exam_filename"] or get_current_exam_name(conn)

    # 获取题目正确答案（优先 exam_questions）
    cursor.execute(
        "SELECT correct_answer, score, type FROM exam_questions WHERE exam_filename = %s AND id = %s",
        (session_exam or "", data.question_id),
    )
    question = cursor.fetchone()
    if not question:
        cursor.execute(
            "SELECT correct_answer, score, type FROM questions WHERE id = %s",
            (data.question_id,),
        )
        question = cursor.fetchone()
    if not question:
        raise HTTPException(status_code=404, detail="题目不存在")

    # 判断答案是否正确
    # 填空题：每个空单独评分
    if question["type"] == "fill":
        fill_blanks = parse_fill_blanks(question["correct_answer"])
        blank_count = len(fill_blanks)
        student_answers = str(data.answer).split("|||")
        total_similarity = 0.0

        if blank_count > 0:
            for i, blank in enumerate(fill_blanks):
                student_answer = student_answers[i] if i < len(student_answers) else ""
                max_similarity = 0.0
                for raw_answer in blank.get("answers", []):
                    similarity = is_fill_answer_match(student_answer, raw_answer)
                    if similarity > max_similarity:
                        max_similarity = similarity
                total_similarity += max_similarity

            score = round(question["score"] * total_similarity / blank_count, 2)
            is_correct = (
                total_similarity >= blank_count - 0.0001
            )  # Allow small floating point error
        else:
            score = 0
            is_correct = False
    else:
        # 选择题：整体判定
        correct_answers = str(question["correct_answer"]).replace("，", ",").split(",")
        student_answer = str(data.answer).strip()
        is_correct = any(
            correct.strip() == student_answer for correct in correct_answers
        )
        score = question["score"] if is_correct else 0

    # 保存或更新答案
    cursor.execute(
        """
        INSERT INTO answers
        (session_id, student_id, question_id, answer, is_correct, score, answer_time, created_at, exam_filename)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT(session_id, question_id) DO UPDATE SET
            answer = excluded.answer,
            is_correct = excluded.is_correct,
            score = excluded.score,
            answer_time = excluded.answer_time,
            created_at = excluded.created_at,
            exam_filename = excluded.exam_filename
        """,
        (
            data.session_id,
            data.student_id,
            data.question_id,
            data.answer,
            is_correct,
            score,
            data.answer_time,
            datetime.now(),
            session_exam or "",
        ),
    )
    conn.commit()

    if get_exam_mode(conn) == "exam":
        return {"score": score}

    # Calculate correct_count as the number of blanks with >= 90% similarity (full score)
    correct_count_value = None
    blank_count_value = None
    if question["type"] == "fill":
        correct_count_value = 0
        if blank_count > 0:
            for i, blank in enumerate(fill_blanks):
                student_answer = student_answers[i] if i < len(student_answers) else ""
                max_similarity = 0.0
                for raw_answer in blank.get("answers", []):
                    similarity = is_fill_answer_match(student_answer, raw_answer)
                    if similarity > max_similarity:
                        max_similarity = similarity
                if max_similarity >= 0.9:
                    correct_count_value += 1
        blank_count_value = blank_count

    return {
        "is_correct": is_correct,
        "score": score,
        "correct_answer": question["correct_answer"],
        "correct_count": correct_count_value,
        "blank_count": blank_count_value,
    }


# ========== 试卷管理接口 ==========


@app.get("/api/exams/list")
def get_exam_list(conn: sqlite3.Connection = Depends(get_db)):
    """获取可用的试卷列表"""
    exams = []
    for exam_path in list_exam_files():
        try:
            data = load_exam_json(exam_path)
            exams.append(
                {
                    "filename": exam_path.name,
                    "title": data.get("title", exam_path.name),
                    "question_count": len(data.get("questions", [])),
                }
            )
        except Exception:
            pass

    return exams


@app.get("/api/exams/current")
def get_current_exam(conn: sqlite3.Connection = Depends(get_db)):
    """获取当前使用的试卷"""
    current_exam = get_current_exam_name(conn)
    return {"current_exam": current_exam}


class ExamSwitch(BaseModel):
    filename: str


class ExamModeSwitch(BaseModel):
    mode: str


class ExamDistributionSwitch(BaseModel):
    distribution: str  # random | fixed
    filename: Optional[str] = None


@app.post("/api/exams/switch")
def switch_exam(data: ExamSwitch, conn: sqlite3.Connection = Depends(get_db)):
    """切换当前使用的试卷"""
    filename = Path(data.filename).name
    exam_path = resolve_exam_path(filename)
    if not exam_path:
        raise HTTPException(status_code=404, detail="试卷文件不存在")

    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM sessions WHERE status = 'active'")
    active_count = _fetchone_value(cursor.fetchone(), 0) or 0
    if active_count > 0:
        cursor.execute(
            "UPDATE sessions SET status = 'abandoned', end_time = %s WHERE status = 'active'",
            (datetime.now(),),
        )

    import_result = import_exam_into_questions(conn, exam_path)
    return {
        "message": "试卷切换成功",
        "current_exam": import_result["filename"],
        "question_count": import_result["question_count"],
        "abandoned_sessions": active_count,
    }


@app.get("/api/exams/mode")
def get_exam_mode_api(conn: sqlite3.Connection = Depends(get_db)):
    return {"mode": get_exam_mode(conn)}


@app.post("/api/exams/mode")
def set_exam_mode_api(data: ExamModeSwitch, conn: sqlite3.Connection = Depends(get_db)):
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM sessions WHERE status = 'active'")
    active_count = _fetchone_value(cursor.fetchone(), 0) or 0
    if active_count > 0:
        cursor.execute(
            "UPDATE sessions SET status = 'abandoned', end_time = %s WHERE status = 'active'",
            (datetime.now(),),
        )
        conn.commit()

    set_exam_mode(conn, data.mode)
    return {"mode": data.mode, "abandoned_sessions": active_count}


@app.get("/api/exams/distribution")
def get_exam_distribution_api(conn: sqlite3.Connection = Depends(get_db)):
    dist, fixed = get_exam_distribution(conn)
    return {"distribution": dist, "filename": fixed}


@app.post("/api/exams/distribution")
def set_exam_distribution_api(
    data: ExamDistributionSwitch, conn: sqlite3.Connection = Depends(get_db)
):
    dist = data.distribution
    filename = data.filename
    if dist == "fixed":
        if not filename:
            raise HTTPException(status_code=400, detail="固定分发必须指定试卷文件名")
        if not resolve_exam_path(filename):
            raise HTTPException(status_code=404, detail="指定试卷不存在")

    set_exam_distribution(conn, dist, filename)
    return {"distribution": dist, "filename": filename or ""}


@app.get("/api/health")
def get_health_status():
    """系统轻量健康检查，用于前端提示教师处理。"""
    return get_system_health()


@app.post("/api/exams/upload")
async def upload_exam(
    file: UploadFile = File(...), conn: sqlite3.Connection = Depends(get_db)
):
    """上传新的试卷 JSON 文件。"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")

    safe_name = Path(file.filename).name
    if not safe_name.lower().endswith(".json"):
        raise HTTPException(status_code=400, detail="仅支持 JSON 试卷文件")

    EXAMS_DIR.mkdir(parents=True, exist_ok=True)
    target_path = EXAMS_DIR / safe_name
    existed = target_path.exists()

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="文件内容为空")

    try:
        data = json.loads(content.decode("utf-8"))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"JSON 解析失败: {e}") from e

    questions = data.get("questions")
    if not isinstance(questions, list) or not questions:
        raise HTTPException(
            status_code=400, detail="试卷格式错误：questions 必须为非空数组"
        )

    with open(target_path, "wb") as f:
        f.write(content)

    # 缓存到 exam_questions，便于多试卷并行练习/考试
    import_exam_into_exam_questions(conn, target_path, data=data)

    current_exam = get_current_exam_name(conn)
    abandoned_sessions = 0
    synced_to_db = False
    if current_exam == safe_name:
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM sessions WHERE status = 'active'")
        abandoned_sessions = _fetchone_value(cursor.fetchone(), 0) or 0
        if abandoned_sessions > 0:
            cursor.execute(
                "UPDATE sessions SET status = 'abandoned', end_time = %s WHERE status = 'active'",
                (datetime.now(),),
            )
            conn.commit()

        # 当前试卷被同名覆盖后，立即重导入题库，保证数据库与文件一致
        import_exam_into_questions(conn, target_path)
        synced_to_db = True

    if existed and synced_to_db:
        message = "同名试卷已覆盖，且已同步更新当前考试题库"
    elif existed:
        message = "同名试卷已覆盖"
    else:
        message = "试卷上传成功"

    return {
        "message": message,
        "filename": target_path.name,
        "title": data.get("title", target_path.name),
        "question_count": len(questions),
        "overwritten": existed,
        "synced_to_db": synced_to_db,
        "abandoned_sessions": abandoned_sessions,
    }


# ========== 统计分析接口（核心功能） ==========


@app.get("/api/analysis/question/{question_id}")
def get_question_analysis(question_id: str, conn: sqlite3.Connection = Depends(get_db)):
    """
    获取某道题目的答题情况分析
    包括：哪些学生答了、答案是什么、是否正确、用时等
    """
    cursor = conn.cursor()
    current_exam = get_current_exam_name(conn) or ""

    # 获取题目信息
    cursor.execute("SELECT * FROM questions WHERE id = %s", (question_id,))
    question = cursor.fetchone()
    if not question:
        raise HTTPException(status_code=404, detail="题目不存在")

    # 获取所有答题记录
    cursor.execute(
        """
        SELECT 
            s.name,
            s.class_number,
            s.student_number,
            a.answer,
            a.is_correct,
            a.score,
            a.answer_time,
            a.created_at
        FROM answers a
        JOIN students s ON a.student_id = s.id
        WHERE a.question_id = %s AND a.exam_filename = %s
        ORDER BY a.created_at DESC
        """,
        (question_id, current_exam),
    )
    answers = [dict(row) for row in cursor.fetchall()]

    # 统计信息
    total = len(answers)
    correct = sum(1 for a in answers if a["is_correct"])
    wrong = total - correct
    correct_rate = round(100.0 * correct / total, 2) if total > 0 else 0

    # 各选项选择人数统计（选择题）
    option_stats = {}
    if question["type"] == "choice":
        for ans in answers:
            opt = ans["answer"]
            option_stats[opt] = option_stats.get(opt, 0) + 1

    return {
        "question": {
            "id": question["id"],
            "type": question["type"],
            "content": question["question"],
            "correct_answer": question["correct_answer"],
        },
        "statistics": {
            "total_attempts": total,
            "correct_count": correct,
            "wrong_count": wrong,
            "correct_rate": correct_rate,
            "option_distribution": option_stats,
        },
        "answers": answers,
    }


@app.get("/api/analysis/student/{student_id}")
def get_student_analysis(
    student_id: int,
    include_attempts: bool = True,
    conn: sqlite3.Connection = Depends(get_db),
):
    """获取某个学生的答题分析"""
    cursor = conn.cursor()
    current_exam = get_current_exam_name(conn) or ""

    # 获取学生信息
    cursor.execute("SELECT * FROM students WHERE id = %s", (student_id,))
    student = cursor.fetchone()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    # 获取答题统计
    cursor.execute(
        """
        SELECT 
            COUNT(*) as total_answered,
            SUM(CASE WHEN is_correct = 1 THEN 1 ELSE 0 END) as correct_count,
            SUM(score) as total_score,
            AVG(answer_time) as avg_time
        FROM answers
        WHERE student_id = %s AND exam_filename = %s
        """,
        (student_id, current_exam),
    )
    stats = dict(cursor.fetchone())

    details = []
    attempts = []

    if include_attempts:
        # 读取该学生该试卷下的所有考试会话（每次考试一条记录）
        cursor.execute(
            """
            SELECT id, start_time, end_time, total_score, status, exam_filename
            FROM sessions
            WHERE student_id = %s AND exam_filename = %s
            ORDER BY start_time ASC
            """,
            (student_id, current_exam),
        )
        session_rows = [dict(row) for row in cursor.fetchall()]

        for idx, session_row in enumerate(session_rows, start=1):
            cursor.execute(
                """
                SELECT 
                    q.id,
                    q.question,
                    q.type,
                    q.correct_answer,
                    a.answer,
                    a.is_correct,
                    a.score,
                    a.answer_time,
                    a.created_at
                FROM answers a
                JOIN questions q ON a.question_id = q.id
                WHERE a.session_id = %s AND a.exam_filename = %s
                ORDER BY a.created_at ASC
                """,
                (session_row["id"], current_exam),
            )
            answer_rows = [dict(row) for row in cursor.fetchall()]

            answered_count = len(answer_rows)
            correct_count = sum(1 for a in answer_rows if a.get("is_correct"))
            avg_time = (
                sum(a.get("answer_time") or 0 for a in answer_rows) / answered_count
                if answered_count > 0
                else 0
            )

            attempts.append(
                {
                    "session_id": session_row["id"],
                    "attempt_no": idx,
                    "exam_filename": session_row.get("exam_filename") or current_exam,
                    "start_time": session_row.get("start_time"),
                    "end_time": session_row.get("end_time"),
                    "status": session_row.get("status"),
                    "total_score": session_row.get("total_score") or 0,
                    "answered_count": answered_count,
                    "correct_count": correct_count,
                    "avg_time": avg_time,
                    "answers": answer_rows,
                }
            )

        # 最新尝试排在前面
        attempts = list(reversed(attempts))

    # 兼容旧前端：保留当前试卷下的答题明细（不分次）
    cursor.execute(
        """
        SELECT 
            q.id,
            q.question,
            q.type,
            q.correct_answer,
            a.answer,
            a.is_correct,
            a.score,
            a.answer_time,
            a.created_at
        FROM answers a
        JOIN questions q ON a.question_id = q.id
        WHERE a.student_id = %s AND a.exam_filename = %s
        ORDER BY a.created_at DESC
        """,
        (student_id, current_exam),
    )
    details = [dict(row) for row in cursor.fetchall()]

    return {
        "student": {
            "id": student["id"],
            "name": student["name"],
            "class": student["class_number"],
            "exam_number": student["exam_number"],
        },
        "statistics": stats,
        "attempts": attempts,
        "answers": details,
    }


@app.get("/api/analysis/overview")
def get_overview(
    exam_filename: Optional[str] = None, conn: sqlite3.Connection = Depends(get_db)
):
    """获取整体统计概览"""
    cursor = conn.cursor()
    current_exam = Path(exam_filename).name if exam_filename else get_current_exam_name(conn) or ""

    # 总参与人数
    cursor.execute(
        "SELECT COUNT(DISTINCT student_id) FROM answers WHERE exam_filename = %s",
        (current_exam,),
    )
    total_students = _fetchone_value(cursor.fetchone(), 0)

    # 总答题次数
    cursor.execute(
        "SELECT COUNT(*) FROM answers WHERE exam_filename = %s",
        (current_exam,),
    )
    total_answers = _fetchone_value(cursor.fetchone(), 0)

    # 正在答题人数
    cursor.execute(
        "SELECT COUNT(*) FROM sessions WHERE status = 'active' AND exam_filename = %s",
        (current_exam,),
    )
    active_sessions = _fetchone_value(cursor.fetchone(), 0) or 0

    # 每道题的统计
    cursor.execute(
        """
        SELECT 
            q.id,
            q.question,
            COUNT(a.id) as attempt_count,
            SUM(CASE WHEN a.is_correct = 1 THEN 1 ELSE 0 END) as correct_count,
            CASE
                WHEN COUNT(a.id) = 0 THEN 0
                ELSE ROUND(100.0 * SUM(CASE WHEN a.is_correct = 1 THEN 1 ELSE 0 END) / COUNT(a.id), 2)
            END as correct_rate
        FROM questions q
        LEFT JOIN answers a ON q.id = a.question_id AND a.exam_filename = %s
        GROUP BY q.id
        ORDER BY correct_rate ASC
        """,
        (current_exam,),
    )
    question_stats = [dict(row) for row in cursor.fetchall()]

    return {
        "total_students": total_students,
        "total_answers": total_answers,
        "active_sessions": active_sessions,
        "question_statistics": question_stats,
    }


@app.get("/api/analysis/active_sessions")
def get_active_sessions(
    exam_filename: Optional[str] = None, conn: sqlite3.Connection = Depends(get_db)
):
    """获取正在答题的学生情况（默认所有试卷，可按试卷过滤）。"""
    cursor = conn.cursor()
    params = []
    where_clause = "s.status = 'active'"
    if exam_filename:
        where_clause += " AND s.exam_filename = %s"
        params.append(exam_filename)

    cursor.execute(
        f"""
        SELECT
            s.id as session_id,
            s.student_id,
            s.exam_filename,
            st.class_number,
            st.student_number,
            st.name,
            COUNT(a.id) as answered_count,
            SUM(CASE WHEN a.is_correct = 1 THEN 1 ELSE 0 END) as correct_count,
            SUM(CASE WHEN a.is_correct = 0 THEN 1 ELSE 0 END) as wrong_count,
            COALESCE(SUM(a.score), 0) as total_score
        FROM sessions s
        JOIN students st ON s.student_id = st.id
        LEFT JOIN answers a ON a.session_id = s.id
        WHERE {where_clause}
        GROUP BY s.id
        ORDER BY st.class_number, st.student_number, s.start_time
        """,
        params,
    )
    rows = cursor.fetchall()
    return [dict(row) for row in rows]


@app.get("/api/analysis/export")
def export_exam_answers(
    exam_filename: Optional[str] = None, conn: sqlite3.Connection = Depends(get_db)
):
    """导出当前试卷答题情况（Excel，每次考试一行，含每题得分）。"""
    current_exam = exam_filename or get_current_exam_name(conn) or ""
    if not current_exam:
        raise HTTPException(status_code=400, detail="当前无可用试卷")

    cursor = conn.cursor()

    # 题目顺序：按类型+序号排序
    cursor.execute("SELECT id, type FROM questions ORDER BY id")
    question_rows = [dict(r) for r in cursor.fetchall()]

    def question_sort_key(q):
        qid = q["id"]
        parts = qid.split("-", 1)
        prefix = parts[0]
        num = 0
        if len(parts) == 2:
            try:
                num = int(parts[1])
            except Exception:
                num = 0
        return (prefix, num, qid)

    question_rows.sort(key=question_sort_key)
    question_ids = [q["id"] for q in question_rows]

    # 计算每位学生在该试卷下的考试次数序号
    cursor.execute(
        """
        SELECT id, student_id, start_time
        FROM sessions
        WHERE exam_filename = %s
        ORDER BY student_id, start_time
        """,
        (current_exam,),
    )
    session_rows = cursor.fetchall()
    attempt_no_by_session = {}
    last_student = None
    count = 0
    for row in session_rows:
        if row["student_id"] != last_student:
            last_student = row["student_id"]
            count = 1
        else:
            count += 1
        attempt_no_by_session[row["id"]] = count

    cursor.execute(
        """
        SELECT 
            s.id as session_id,
            s.student_id,
            s.start_time,
            s.end_time,
            s.total_score,
            s.status,
            s.exam_filename,
            st.class_number,
            st.student_number,
            st.name
        FROM sessions s
        JOIN students st ON s.student_id = st.id
        WHERE s.exam_filename = %s
        ORDER BY st.class_number, st.student_number, s.start_time
        """,
        (current_exam,),
    )
    sessions = [dict(r) for r in cursor.fetchall()]

    # 取所有答题得分
    cursor.execute(
        """
        SELECT session_id, question_id, score
        FROM answers
        WHERE exam_filename = %s
        """,
        (current_exam,),
    )
    answer_rows = cursor.fetchall()
    scores_by_session = {}
    for row in answer_rows:
        session_id = row["session_id"]
        scores_by_session.setdefault(session_id, {})[row["question_id"]] = (
            row["score"] or 0
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "答题情况"

    headers = [
        "班级",
        "学号",
        "姓名",
        "总分",
        "试卷",
        "考试时间",
        "第几次考试",
    ]
    for idx in range(1, len(question_ids) + 1):
        headers.append(str(idx))
    ws.append(headers)

    for session in sessions:
        session_id = session["session_id"]
        attempt_no = attempt_no_by_session.get(session_id, 1)
        start_time = session.get("start_time")
        end_time = session.get("end_time")
        exam_time = f"{start_time} - {end_time}" if end_time else f"{start_time}"
        scores_map = scores_by_session.get(session_id, {})
        row_values = [
            session["class_number"],
            session["student_number"],
            session["name"],
            session.get("total_score") or 0,
            session.get("exam_filename") or current_exam,
            exam_time,
            attempt_no,
        ]
        for qid in question_ids:
            row_values.append(scores_map.get(qid, 0))
        ws.append(row_values)

    # 简单设置列宽，避免乱码/过窄
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["C"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{current_exam}_答题情况.xlsx"
    quoted_name = quote(filename)
    headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{quoted_name}"}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


# 静态文件服务
@app.get("/")
def root():
    return RedirectResponse(url="/xx", status_code=307)


@app.get("/xx")
@app.get("/xx/")
def student_entry():
    return FileResponse(APP_DIR / "index.html", headers={"Cache-Control": "no-store"})


@app.get("/xx/js")
@app.get("/xx/js/")
def teacher_entry():
    return FileResponse(APP_DIR / "teacher.html", headers={"Cache-Control": "no-store"})


@app.get("/xx/student_detail.html")
def teacher_student_detail():
    return FileResponse(
        APP_DIR / "student_detail.html", headers={"Cache-Control": "no-store"}
    )


@app.get("/styles.css")
def serve_styles():
    return FileResponse(BASE_DIR / "styles.css", headers={"Cache-Control": "no-store"})


@app.get("/xx/styles.css")
def serve_styles_xx():
    return FileResponse(BASE_DIR / "styles.css", headers={"Cache-Control": "no-store"})


@app.get("/app/{path:path}")
def serve_app(path: str):
    file_path = (APP_DIR / path).resolve()
    if APP_DIR.resolve() in file_path.parents and file_path.exists():
        headers = {"Cache-Control": "no-store"} if file_path.suffix == ".html" else None
        return FileResponse(file_path, headers=headers)
    return {"detail": "Not Found"}


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
